# pharmacy_app.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime, timedelta
import json
import os
import csv
import hashlib
import uuid
import platform
import shutil
from pathlib import Path
import matplotlib
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
import numpy as np
import sys

class PharmacyApp:
    def __init__(self, root):
        self.root = root
        self.root.title("MERawi Pharmacy Pro - Complete Pharmacy Management System")
        self.root.geometry("1200x700")
        
        # Initialize data directories FIRST
        self.ensure_data_directories()
        
        # Configure style
        self.setup_styles()
        
        # Create database and tables
        self.create_database()
        
        # Create menu bar
        self.create_menu()
        
        # Create status bar
        self.create_status_bar()
        
        # Create main content area with notebook
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Create tabs
        self.create_medicines_tab()
        self.create_sales_tab()
        self.create_reports_tab()
        self.create_dashboard_tab()
        self.create_bin_card_tab()
        
        # Load initial data
        self.load_medicines()
        self.check_alerts()
        
        self._last_tab = self.medicines_frame
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
        self._switching_tab = False
    
    def show_settings(self):
        config = self.load_config()
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Settings")
        dialog.geometry("400x250")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Mode selection
        mode_frame = tk.LabelFrame(dialog, text="Sales Mode", padx=10, pady=10)
        mode_frame.pack(fill='x', padx=10, pady=5)
        
        mode_var = tk.BooleanVar(value=config.get('government_mode', False))
        tk.Checkbutton(mode_frame, text="Government Mode (enables TIN and fiscal printer)", 
                    variable=mode_var).pack(anchor='w')
        
        # TIN entry
        tin_frame = tk.LabelFrame(dialog, text="Taxpayer Identification Number (TIN)", padx=10, pady=10)
        tin_frame.pack(fill='x', padx=10, pady=5)
        
        tin_entry = tk.Entry(tin_frame, width=30)
        tin_entry.insert(0, config.get('tin', ''))
        tin_entry.pack(anchor='w', pady=5)
        
        # Function to show/hide TIN field based on mode
        def toggle_tin_state(*args):
            if mode_var.get():
                tin_entry.config(state='normal')
            else:
                tin_entry.config(state='disabled')
                tin_entry.delete(0, tk.END)
        mode_var.trace_add('write', toggle_tin_state)
        toggle_tin_state()   # initial state
        
        # Save button
        def save_settings():
            new_config = {
                'government_mode': mode_var.get(),
                'tin': tin_entry.get().strip() if mode_var.get() else ''
            }
            self.save_config(new_config)
            self.apply_settings()
            messagebox.showinfo("Settings Saved", "The sales tab has been updated.")
            dialog.destroy()
        
        tk.Button(dialog, text="Save", command=save_settings, bg='#2c3e50', fg='white',
                font=('Segoe UI', 10), padx=10).pack(pady=10)
    
    def load_config(self):
        config_file = self.get_config_path()
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r') as f:
                    return json.load(f)
            except:
                return {}
        return {}

    def save_config(self, config):
        with open(self.get_config_path(), 'w') as f:
            json.dump(config, f, indent=4)
    
    def get_config_path(self):
        return os.path.join(self.get_app_data_dir(), 'config.json')
        
    # ========== DATA DIRECTORY MANAGEMENT ==========
    
    def get_app_data_dir(self):
        """Get the appropriate application data directory for all platforms"""
        system = platform.system()
        
        if system == "Windows":
            base_dir = os.path.join(os.environ['LOCALAPPDATA'], 'MERawiPharmacy')
        elif system == "Darwin":  # macOS
            base_dir = os.path.join(str(Path.home()), 'Library', 'Application Support', 'MERawiPharmacy')
        else:  # Linux
            base_dir = os.path.join(str(Path.home()), '.local', 'share', 'MERawiPharmacy')
        
        os.makedirs(base_dir, exist_ok=True)
        return base_dir
    
    def get_database_path(self):
        """Get the database file path"""
        return os.path.join(self.get_app_data_dir(), 'pharmacy.db')
    
    def get_license_path(self):
        return os.path.join(self.get_app_data_dir(), 'license.json')
    
    def get_backups_dir(self):
        """Get the backups directory"""
        backups_dir = os.path.join(self.get_app_data_dir(), 'backups')
        os.makedirs(backups_dir, exist_ok=True)
        return backups_dir
    
    def get_receipts_dir(self):
        """Get the receipts directory"""
        receipts_dir = os.path.join(self.get_app_data_dir(), 'receipts')
        os.makedirs(receipts_dir, exist_ok=True)
        return receipts_dir
    
    def get_exports_dir(self):
        """Get the exports directory"""
        exports_dir = os.path.join(self.get_app_data_dir(), 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        return exports_dir
    
    def ensure_data_directories(self):
        """Create all necessary data directories"""
        self.get_app_data_dir()
        self.get_backups_dir()
        self.get_receipts_dir()
        self.get_exports_dir()
    
    def show_data_location(self):
        """Show where data is stored"""
        data_dir = self.get_app_data_dir()
        
        message = f"📍 DATA STORAGE LOCATION\n\n"
        message += f"All your pharmacy data is stored at:\n{data_dir}\n\n"
        message += f"Contents:\n"
        message += f"• pharmacy.db - Main database\n"
        message += f"• license.key - License activation\n"
        message += f"• backups/ - Excel backup files\n"
        message += f"• receipts/ - Saved receipts\n"
        message += f"• exports/ - Imported files\n\n"
        message += f"This location is NOT affected when you update the app.\n"
        message += f"Your data is safe even if you uninstall the app!"
        
        if messagebox.askyesno("📍 Data Location", message + "\n\nOpen this folder?"):
            if platform.system() == "Windows":
                os.startfile(data_dir)
            else:
                os.system(f'open "{data_dir}"')
    
    # ========== LICENSE MANAGEMENT ==========
    
    def check_license(self):
        license_file = self.get_license_path()
        
        if not os.path.exists(license_file):
            return self.show_trial_or_purchase_dialog()
        
        # Load license data
        with open(license_file, 'r') as f:
            license_data = json.load(f)
        
        # Paid license -> always OK
        if license_data.get('type') == 'paid':
            return True
        
        # Trial license -> check age
        if license_data.get('type') == 'trial':
            install_date = datetime.strptime(license_data['install_date'], '%Y-%m-%d').date()
            days_used = (datetime.now().date() - install_date).days
            if days_used < 30:
                return True
            else:
                # Trial expired : delete license file
                os.remove(license_file)
                return self.show_trial_or_purchase_dialog()
        
        # Unknown type -> ask again
        return self.show_trial_or_purchase_dialog()

    def show_trial_or_purchase_dialog(self):
        """Professional purchase dialog with scrollbar + trial option"""
        # Create dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("License Activation")
        dialog.geometry("550x650")
        dialog.configure(bg='#f5f5f5')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Center window
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (550 // 2)
        y = (dialog.winfo_screenheight() // 2) - (650 // 2)
        dialog.geometry(f'+{x}+{y}')
        
        # Bring to front
        dialog.lift()
        dialog.focus_force()
        
        # Variable to track activation
        activated = tk.BooleanVar(value=False)
        
        # Create canvas and scrollbar
        canvas = tk.Canvas(dialog, bg='#f5f5f5', highlightthickness=0)
        scrollbar = tk.Scrollbar(dialog, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f5f5f5')
        
        scrollable_frame.bind(
            '<Configure>',
            lambda e: canvas.configure(scrollregion=canvas.bbox('all'))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Mouse wheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), 'units')
        
        scrollable_frame.bind('<Enter>', lambda e: canvas.bind_all('<MouseWheel>', _on_mousewheel))
        scrollable_frame.bind('<Leave>', lambda e: canvas.unbind_all('<MouseWheel>'))
        
        # Pack canvas and scrollbar
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Main content frame
        main_frame = tk.Frame(scrollable_frame, bg='white', padx=30, pady=30)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Header
        tk.Label(main_frame, text="🔒", bg='white', font=('Segoe UI', 40)).pack(pady=(0, 10))
        tk.Label(main_frame, text="MERawi PHARMACY PRO", bg='white', fg='#2c3e50',
                font=('Segoe UI', 18, 'bold')).pack()
        tk.Label(main_frame, text="License Required", bg='white', fg='#e67e22',
                font=('Segoe UI', 12)).pack(pady=(5, 20))
        
        # Computer ID
        id_frame = tk.Frame(main_frame, bg='#f8f9fa', padx=15, pady=15)
        id_frame.pack(fill='x', pady=10)
        
        tk.Label(id_frame, text="YOUR COMPUTER ID:", bg='#f8f9fa', fg='#2c3e50',
                font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        
        computer_id = self.get_pc_fingerprint()
        
        id_display = tk.Entry(id_frame, font=('Courier', 11), bg='white', fg='#2c3e50',
                            relief='solid', bd=1)
        id_display.insert(0, computer_id)
        id_display.config(state='readonly')
        id_display.pack(fill='x', pady=5)
        
        def copy_id():
            dialog.clipboard_clear()
            dialog.clipboard_append(computer_id)
            copy_btn.config(text="✓ Copied!", bg='#27ae60')
            dialog.after(1500, lambda: copy_btn.config(text="📋 Copy Computer ID", bg='#3498db'))
        
        copy_btn = tk.Button(id_frame, text="📋 Copy Computer ID", bg='#3498db', fg='white',
                            font=('Segoe UI', 9), relief='flat', command=copy_id)
        copy_btn.pack(pady=5)
        
        tk.Label(id_frame, text="Send this ID to get your license key", bg='#f8f9fa', fg='#7f8c8d',
                font=('Segoe UI', 8, 'italic')).pack()
        
        # Contact information
        contact_frame = tk.Frame(main_frame, bg='#f8f9fa', padx=15, pady=15)
        contact_frame.pack(fill='x', pady=10)
        
        tk.Label(contact_frame, text="CONTACT FOR LICENSE", bg='#f8f9fa', fg='#2c3e50',
                font=('Segoe UI', 10, 'bold')).pack(pady=(0, 10))
        
        contacts = [
            ("Developer:", "Merawi Yohannes"),
            ("Phone 1:", "0921-540-245"),
            ("Phone 2:", "0960-633-549"),
            ("Email:", "merawiyohannes@gmail.com"),
        ]
        
        for label, value in contacts:
            frame = tk.Frame(contact_frame, bg='#f8f9fa')
            frame.pack(fill='x', pady=2)
            tk.Label(frame, text=label, bg='#f8f9fa', fg='#7f8c8d',
                    font=('Segoe UI', 9), width=10, anchor='w').pack(side='left')
            tk.Label(frame, text=value, bg='#f8f9fa', fg='#2c3e50',
                    font=('Segoe UI', 9)).pack(side='left')
        
        # Price
        price_frame = tk.Frame(main_frame, bg='#2c3e50', padx=15, pady=10)
        price_frame.pack(fill='x', pady=10)
        
        tk.Label(price_frame, text="20k - 30k ETB", bg='#2c3e50', fg='white',
                font=('Segoe UI', 18, 'bold')).pack()
        tk.Label(price_frame, text="One-time payment · Lifetime license", bg='#2c3e50', fg='#ecf0f1',
                font=('Segoe UI', 9)).pack()
        
        # License entry
        entry_frame = tk.Frame(main_frame, bg='white', padx=15, pady=15)
        entry_frame.pack(fill='x', pady=10)
        
        tk.Label(entry_frame, text="ENTER LICENSE KEY:", bg='white', fg='#2c3e50',
                font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        
        key_entry = tk.Entry(entry_frame, width=25, font=('Courier', 11),
                            justify='center', relief='solid', bd=1)
        key_entry.pack(pady=5, ipady=5)
        key_entry.focus_set()
        
        def activate():
            key = key_entry.get().strip().upper()
            expected_key = self.generate_license_key(computer_id)
            
            if key == expected_key:
                # Store as paid license (JSON format)
                license_data = {"type": "paid"}
                with open(self.get_license_path(), 'w') as f:
                    json.dump(license_data, f)
                messagebox.showinfo("✅ Success", "License activated successfully!\n\nYour data is stored in:\n" + self.get_app_data_dir())
                activated.set(True)
                dialog.destroy()
            else:
                messagebox.showerror("❌ Error", "Invalid license key.\n\nPlease check and try again.")
                key_entry.delete(0, tk.END)
                key_entry.focus_set()
        
        def close_dialog():
            dialog.destroy()
            activated.set(False)
        
        def start_trial():
            # 1) Check if trial was already used on this PC (simple flag file)
            trial_flag = os.path.join(self.get_app_data_dir(), '.trial_used')
            if os.path.exists(trial_flag):
                with open(trial_flag, 'r') as f:
                    stored_id = f.read().strip()
                if stored_id == computer_id:   # computer_id is already defined above
                    messagebox.showerror(
                        "Trial Unavailable",
                        "Free trial already used on this computer.\n"
                        "Please purchase a license (use Activate button)."
                    )
                    return

            # 2) Validate the license key (same as Activate button)
            key = key_entry.get().strip().upper()
            expected_key = self.generate_license_key(computer_id)
            if key != expected_key:
                messagebox.showerror("❌ Error", "Invalid license key.")
                key_entry.delete(0, tk.END)
                key_entry.focus_set()
                return

            # 3) Store trial license with timestamp
            trial_data = {
                "type": "trial",
                "install_date": datetime.now().strftime("%Y-%m-%d")
            }
            with open(self.get_license_path(), 'w') as f:
                json.dump(trial_data, f)

            # 4) Mark trial as used (so user cannot start another trial later)
            with open(trial_flag, 'w') as f:
                f.write(computer_id)

            messagebox.showinfo(
                "Trial Started",
                "30-day free trial activated! App will now restart.\n"
                "After 30 days the license will expire and you will need to Activate."
            )
            dialog.destroy()
            self.root.destroy()
            os.execv(sys.executable, [sys.executable] + sys.argv)
        
        # Buttons frame
        btn_frame = tk.Frame(main_frame, bg='white')
        btn_frame.pack(fill='x', pady=15)
        
        tk.Button(btn_frame, text="Activate License", bg='#2c3e50', fg='white',
                font=('Segoe UI', 10), padx=20, pady=8, relief='flat',
                command=activate).pack(side='left', padx=5)
        
        tk.Button(btn_frame, text="Exit", bg='#95a5a6', fg='white',
                font=('Segoe UI', 10), padx=20, pady=8, relief='flat',
                command=close_dialog).pack(side='left', padx=5)
        
        # NEW: Trial button
        tk.Button(btn_frame, text="Start 30-Day Free Trial", bg='#27ae60', fg='white',
                font=('Segoe UI', 10), padx=20, pady=8, relief='flat',
                command=start_trial).pack(side='left', padx=5)
        
        # Handle window close button
        dialog.protocol("WM_DELETE_WINDOW", close_dialog)
        
        # Wait for dialog to close
        self.root.wait_window(dialog)
        
        return activated.get()

    def get_pc_fingerprint(self):
        """Get unique computer ID"""
        fingerprint = str(uuid.getnode())
        return hashlib.md5(fingerprint.encode()).hexdigest()[:16].upper()
    
    def generate_license_key(self, computer_id):
        """Generate license key from computer ID"""
        secret = "MERAWI2026PRO"
        combined = computer_id + secret
        key = hashlib.md5(combined.encode()).hexdigest()[:16].upper()
        return f"{key[:4]}-{key[4:8]}-{key[8:12]}-{key[12:16]}"


    # ========== DATABASE MANAGEMENT ==========
    
    def view_invoice_details(self, invoice_id):
        messagebox.showinfo("Coming Soon", "Payment recording will be implemented next.")
    
    def show_invoice_analysis(self):
        """Invoice management window – uses SQL for overdue detection (consistent with alerts)"""
        import datetime
        
        conn = sqlite3.connect(self.get_database_path())
        cursor = conn.cursor()
        
        # Get all invoices
        cursor.execute('''
            SELECT id, invoice_number, customer_name, total, amount_paid, due_date, status 
            FROM invoices 
            ORDER BY date_issued DESC
        ''')
        all_invoices = cursor.fetchall()
        
        # Get the set of overdue invoice IDs (same SQL as check_alerts)
        cursor.execute('''
            SELECT id FROM invoices 
            WHERE status IN ('Pending', 'Partial') AND due_date < date('now')
        ''')
        overdue_ids = {row[0] for row in cursor.fetchall()}
        
        conn.close()
        
        # Calculate totals
        total_outstanding = 0
        total_overdue = 0
        total_paid = 0
        for inv in all_invoices:
            inv_id, _, _, total, paid, _, status = inv
            if status in ('Pending', 'Partial'):
                due = total - paid
                total_outstanding += due
                if inv_id in overdue_ids:
                    total_overdue += due
            total_paid += paid
        
        # ----- GUI -----
        win = tk.Toplevel(self.root)
        win.title("📊 Invoice Analysis")
        win.geometry("1000x700")
        win.transient(self.root)
        
        # Summary bar
        summary = tk.Frame(win, bg='#2c3e50', padx=20, pady=15)
        summary.pack(fill='x')
        tk.Label(summary, text="Total Outstanding:", bg='#2c3e50', fg='white', font=('Segoe UI', 11)).pack(side='left', padx=15)
        tk.Label(summary, text=f"{total_outstanding:,.2f} Birr", bg='#2c3e50', fg='#f1c40f', font=('Segoe UI', 13, 'bold')).pack(side='left', padx=5)
        tk.Label(summary, text="Overdue:", bg='#2c3e50', fg='white', font=('Segoe UI', 11)).pack(side='left', padx=15)
        tk.Label(summary, text=f"{total_overdue:,.2f} Birr", bg='#2c3e50', fg='#e74c3c', font=('Segoe UI', 13, 'bold')).pack(side='left', padx=5)
        tk.Label(summary, text="Total Paid:", bg='#2c3e50', fg='white', font=('Segoe UI', 11)).pack(side='left', padx=15)
        tk.Label(summary, text=f"{total_paid:,.2f} Birr", bg='#2c3e50', fg='#2ecc71', font=('Segoe UI', 13, 'bold')).pack(side='left', padx=5)
        
        # Filter row
        filter_frame = tk.Frame(win, bg='#ecf0f1', padx=10, pady=5)
        filter_frame.pack(fill='x')
        show_overdue_only = tk.BooleanVar(value=False)
        
        # Treeview
        tree_frame = tk.Frame(win)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        columns = ('ID', 'Invoice #', 'Customer', 'Total', 'Paid', 'Due Date', 'Status')
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=18)
        widths = [50, 120, 220, 100, 100, 100, 100]
        for i, col in enumerate(columns):
            tree.heading(col, text=col)
            tree.column(col, width=widths[i])
        
        # Color tags
        tree.tag_configure('Pending', background='#fff3cd')
        tree.tag_configure('Partial', background='#ffe6b3')
        tree.tag_configure('Paid', background='#d4edda')
        tree.tag_configure('Overdue', background='#f8d7da')
        
        scroll = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        tree.pack(side='left', fill='both', expand=True)
        scroll.pack(side='right', fill='y')
        
        def refresh_tree():
            for item in tree.get_children():
                tree.delete(item)
            for inv in all_invoices:
                inv_id = inv[0]
                is_overdue = inv_id in overdue_ids
                if show_overdue_only.get() and not is_overdue:
                    continue
                tag = 'Overdue' if is_overdue else inv[6]
                tree.insert('', 'end', values=inv, tags=(tag,))
        
        # Filter checkbox
        tk.Checkbutton(filter_frame, text="Show only overdue invoices", variable=show_overdue_only,
                    command=refresh_tree, bg='#ecf0f1', font=('Segoe UI', 10)).pack(side='left')
        tk.Label(filter_frame, text=f"({len(overdue_ids)} overdue)", bg='#ecf0f1', fg='#e74c3c',
                font=('Segoe UI', 9, 'bold')).pack(side='left', padx=10)
        
        # Buttons
        btn_frame = tk.Frame(win)
        btn_frame.pack(side='bottom', pady=10)
        
        def view_invoice():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Select Invoice", "Please select an invoice.")
                return
            inv_id = tree.item(sel[0])['values'][0]
            # Placeholder – implement payment dialog later
            messagebox.showinfo("Coming Soon", f"Invoice {inv_id} details & payment recording will be added next.")
        
        def reprint_invoice():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Select Invoice", "Please select an invoice.")
                return
            inv_id = tree.item(sel[0])['values'][0]
            self._reprint_invoice_by_id(inv_id)
        
        tk.Button(btn_frame, text="View / Pay", command=view_invoice, bg='#3498db', fg='white', padx=15, pady=5).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Reprint Invoice", command=reprint_invoice, bg='#e67e22', fg='white', padx=15, pady=5).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Refresh", command=lambda: [win.destroy(), self.show_invoice_analysis()], bg='#2c3e50', fg='white', padx=15, pady=5).pack(side='left', padx=5)
        tk.Button(btn_frame, text="Close", command=win.destroy, bg='#95a5a6', fg='white', padx=15, pady=5).pack(side='left', padx=5)
        
        refresh_tree()
                    
    def print_invoice_pdf(self, pdf_data):
        """Print invoice PDF directly (same layout as save)"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            import tempfile
            import subprocess
            import platform
            
            # Create temporary PDF
            fd, temp_path = tempfile.mkstemp(suffix='.pdf')
            os.close(fd)
            
            doc = SimpleDocTemplate(temp_path, pagesize=A4,
                                rightMargin=72, leftMargin=72,
                                topMargin=72, bottomMargin=18)
            
            styles = getSampleStyleSheet()
            story = []
            
            # Header
            header_style = ParagraphStyle('CustomHeader', parent=styles['Heading1'],
                                        fontSize=16, textColor=colors.HexColor('#2c3e50'),
                                        spaceAfter=10, alignment=1)
            story.append(Paragraph("MERawi PHARMACY PRO", header_style))
            story.append(Paragraph("TAX INVOICE", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            # Invoice details
            info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=10, spaceAfter=4)
            story.append(Paragraph(f"Invoice No: {pdf_data['invoice_no']}", info_style))
            story.append(Paragraph(f"Date: {pdf_data['date']}", info_style))
            story.append(Paragraph(f"Customer: {pdf_data['customer_name']}", info_style))
            if pdf_data.get('customer_phone'):
                story.append(Paragraph(f"Phone: {pdf_data['customer_phone']}", info_style))
            if pdf_data.get('customer_address'):
                story.append(Paragraph(f"Address: {pdf_data['customer_address']}", info_style))
            story.append(Paragraph(f"Due Date: {pdf_data['due_date']}", info_style))
            story.append(Spacer(1, 12))
            
            # TIN information
            if pdf_data.get('business_tin') or pdf_data.get('customer_tin'):
                tin_style = ParagraphStyle('TIN', parent=styles['Normal'], fontSize=9, alignment=0)
                if pdf_data.get('business_tin'):
                    story.append(Paragraph(f"Business TIN: {pdf_data['business_tin']}", tin_style))
                if pdf_data.get('customer_tin'):
                    story.append(Paragraph(f"Customer TIN: {pdf_data['customer_tin']}", tin_style))
                story.append(Spacer(1, 10))
            
            # Items table
            data = [['Item', 'Qty', 'Price (Birr)', 'Total (Birr)']]
            for item in pdf_data['items']:
                data.append([
                    item['name'][:40],
                    str(item['qty']),
                    f"{item['price']:.2f}",
                    f"{item['total']:.2f}"
                ])
            
            table = Table(data, colWidths=[3.2*inch, 0.8*inch, 1*inch, 1.2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('ALIGN', (0,0), (0,-1), 'LEFT'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ]))
            story.append(table)
            story.append(Spacer(1, 12))
            
            # Totals
            total_style = ParagraphStyle('Total', parent=styles['Normal'], fontSize=11, alignment=2)
            story.append(Paragraph(f"Subtotal: {pdf_data['subtotal']:.2f} Birr", total_style))
            story.append(Paragraph(f"VAT ({pdf_data['tax_rate']}%): {pdf_data['tax_amount']:.2f} Birr", total_style))
            story.append(Paragraph(f"<b>TOTAL: {pdf_data['total']:.2f} Birr</b>", total_style))
            story.append(Spacer(1, 20))
            
            # Footer
            footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=1, textColor=colors.grey)
            story.append(Paragraph("Thank you for your business!", footer_style))
            story.append(Paragraph("MERawi Yohannes | 0921-540-245 | merawiyohannes@gmail.com", footer_style))
            
            doc.build(story)
            
            # Send to printer
            if platform.system() == "Windows":
                os.startfile(temp_path, "print")
            elif platform.system() == "Darwin":
                subprocess.run(['lp', temp_path])
            else:
                subprocess.run(['lp', temp_path])
            
            messagebox.showinfo("Print", "Invoice sent to printer.")
            
            # Delete temp file after 5 seconds
            self.root.after(5000, lambda: os.unlink(temp_path) if os.path.exists(temp_path) else None)
            
        except ImportError:
            messagebox.showwarning("PDF Library Missing", "ReportLab not installed.\nRun: pip install reportlab")
        except Exception as e:
            messagebox.showerror("Print Error", f"Could not print: {str(e)}")
                
    def save_invoice_pdf(self, pdf_data):
        """Save invoice as PDF file (includes TIN if present)"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            receipts_dir = self.get_receipts_dir()
            filename = f"invoice_{pdf_data['invoice_no']}.pdf"
            filepath = os.path.join(receipts_dir, filename)
            
            doc = SimpleDocTemplate(filepath, pagesize=A4,
                                rightMargin=72, leftMargin=72,
                                topMargin=72, bottomMargin=18)
            
            styles = getSampleStyleSheet()
            story = []
            
            # Header
            header_style = ParagraphStyle('CustomHeader', parent=styles['Heading1'],
                                        fontSize=16, textColor=colors.HexColor('#2c3e50'),
                                        spaceAfter=10, alignment=1)
            story.append(Paragraph("MERawi PHARMACY PRO", header_style))
            story.append(Paragraph("TAX INVOICE", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            # Invoice details
            info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=10, spaceAfter=4)
            story.append(Paragraph(f"Invoice No: {pdf_data['invoice_no']}", info_style))
            story.append(Paragraph(f"Date: {pdf_data['date']}", info_style))
            story.append(Paragraph(f"Customer: {pdf_data['customer_name']}", info_style))
            if pdf_data.get('customer_phone'):
                story.append(Paragraph(f"Phone: {pdf_data['customer_phone']}", info_style))
            if pdf_data.get('customer_address'):
                story.append(Paragraph(f"Address: {pdf_data['customer_address']}", info_style))
            story.append(Paragraph(f"Due Date: {pdf_data['due_date']}", info_style))
            story.append(Spacer(1, 12))
            
            # TIN information (if present)
            if pdf_data.get('business_tin') or pdf_data.get('customer_tin'):
                tin_style = ParagraphStyle('TIN', parent=styles['Normal'], fontSize=9, alignment=0)
                if pdf_data.get('business_tin'):
                    story.append(Paragraph(f"Business TIN: {pdf_data['business_tin']}", tin_style))
                if pdf_data.get('customer_tin'):
                    story.append(Paragraph(f"Customer TIN: {pdf_data['customer_tin']}", tin_style))
                story.append(Spacer(1, 10))
            
            # Items table
            data = [['Item', 'Qty', 'Price (Birr)', 'Total (Birr)']]
            for item in pdf_data['items']:
                data.append([
                    item['name'][:40],
                    str(item['qty']),
                    f"{item['price']:.2f}",
                    f"{item['total']:.2f}"
                ])
            
            table = Table(data, colWidths=[3.2*inch, 0.8*inch, 1*inch, 1.2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('ALIGN', (0,0), (0,-1), 'LEFT'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ]))
            story.append(table)
            story.append(Spacer(1, 12))
            
            # Totals
            total_style = ParagraphStyle('Total', parent=styles['Normal'], fontSize=11, alignment=2)
            story.append(Paragraph(f"Subtotal: {pdf_data['subtotal']:.2f} Birr", total_style))
            story.append(Paragraph(f"VAT ({pdf_data['tax_rate']}%): {pdf_data['tax_amount']:.2f} Birr", total_style))
            story.append(Paragraph(f"<b>TOTAL: {pdf_data['total']:.2f} Birr</b>", total_style))
            story.append(Spacer(1, 20))
            
            # Footer
            footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=1, textColor=colors.grey)
            story.append(Paragraph("Thank you for your business!", footer_style))
            story.append(Paragraph("MERawi Yohannes | 0921-540-245 | merawiyohannes@gmail.com", footer_style))
            
            doc.build(story)
            
            messagebox.showinfo("✅ Invoice Saved", f"Invoice saved as PDF:\n{filepath}")
            
            if messagebox.askyesno("Open Folder", "Open invoices folder?"):
                if platform.system() == "Windows":
                    os.startfile(receipts_dir)
                else:
                    os.system(f'open "{receipts_dir}"')
                    
        except ImportError:
            messagebox.showwarning("PDF Library Missing", "ReportLab not installed.\nRun: pip install reportlab")
        except Exception as e:
            messagebox.showerror("Error", f"Could not save invoice PDF: {str(e)}")
            
    def create_invoice(self):
        """Create invoice – auto‑save PDF, then ask to print"""
        if not self.bill_items:
            messagebox.showwarning("Warning", "No items to invoice")
            return False
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Customer Information")
        dialog.geometry("450x450")
        dialog.transient(self.root)
        dialog.grab_set()
        
        tk.Label(dialog, text="Invoice Customer Details", font=('Segoe UI', 12, 'bold')).pack(pady=10)
        
        frame = tk.Frame(dialog)
        frame.pack(padx=20, pady=5)
        
        row = 0
        tk.Label(frame, text="Name:*").grid(row=row, column=0, sticky='w', pady=5)
        name_entry = tk.Entry(frame, width=35)
        name_entry.grid(row=row, column=1, pady=5)
        row += 1
        
        tk.Label(frame, text="Phone:").grid(row=row, column=0, sticky='w', pady=5)
        phone_entry = tk.Entry(frame, width=35)
        phone_entry.grid(row=row, column=1, pady=5)
        row += 1
        
        tk.Label(frame, text="Address:").grid(row=row, column=0, sticky='w', pady=5)
        addr_entry = tk.Entry(frame, width=35)
        addr_entry.grid(row=row, column=1, pady=5)
        row += 1
        
        # TIN fields if government mode
        customer_tin_entry = None
        if self.government_mode:
            tk.Label(frame, text="Business TIN:").grid(row=row, column=0, sticky='w', pady=5)
            biz_tin_var = tk.StringVar(value=self.business_tin)
            biz_tin_display = tk.Entry(frame, width=35, textvariable=biz_tin_var, state='readonly')
            biz_tin_display.grid(row=row, column=1, pady=5)
            row += 1
            
            tk.Label(frame, text="Customer TIN (optional):").grid(row=row, column=0, sticky='w', pady=5)
            customer_tin_entry = tk.Entry(frame, width=35)
            customer_tin_entry.grid(row=row, column=1, pady=5)
            row += 1
        
        tk.Label(frame, text="Due Date (YYYY-MM-DD):").grid(row=row, column=0, sticky='w', pady=5)
        due_entry = tk.Entry(frame, width=35)
        due_entry.insert(0, (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d'))
        due_entry.grid(row=row, column=1, pady=5)
        
        customer_data = {}
        
        def confirm():
            if not name_entry.get().strip():
                messagebox.showerror("Error", "Customer name is required")
                return
            customer_data['name'] = name_entry.get().strip()
            customer_data['phone'] = phone_entry.get().strip()
            customer_data['address'] = addr_entry.get().strip()
            customer_data['due_date'] = due_entry.get().strip()
            if self.government_mode and customer_tin_entry:
                customer_data['business_tin'] = self.business_tin
                customer_data['customer_tin'] = customer_tin_entry.get().strip()
            else:
                customer_data['business_tin'] = ''
                customer_data['customer_tin'] = ''
            dialog.destroy()
        
        tk.Button(dialog, text="Create Invoice", bg='#27ae60', fg='white', command=confirm).pack(pady=20)
        dialog.wait_window()
        
        if not customer_data:
            return False
        
        # Calculate totals
        subtotal = sum(item['total'] for item in self.bill_items)
        tax_rate = float(self.tax_var.get())
        tax_amount = subtotal * (tax_rate/100)
        total = subtotal + tax_amount
        invoice_no = f"INV-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        try:
            conn = sqlite3.connect(self.get_database_path())
            cursor = conn.cursor()
            
            # Build notes field for TINs
            notes = ""
            if customer_data.get('business_tin'):
                notes += f"Business TIN: {customer_data['business_tin']}\n"
            if customer_data.get('customer_tin'):
                notes += f"Customer TIN: {customer_data['customer_tin']}"
            
            cursor.execute('''
                INSERT INTO invoices (invoice_number, customer_name, customer_phone, customer_address,
                                    date_issued, due_date, subtotal, tax_rate, tax_amount, total, status, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Pending', ?)
            ''', (invoice_no, customer_data['name'], customer_data['phone'], customer_data['address'],
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'), customer_data['due_date'],
                subtotal, tax_rate, tax_amount, total, notes.strip()))
            
            invoice_id = cursor.lastrowid
            
            for item in self.bill_items:
                cursor.execute('SELECT quantity FROM medicines WHERE id = ?', (item['id'],))
                old_qty = cursor.fetchone()[0]
                new_qty = old_qty - item['qty']
                cursor.execute('UPDATE medicines SET quantity = ? WHERE id = ?', (new_qty, item['id']))
                cursor.execute('''
                    INSERT INTO stock_movements (medicine_id, date, type, quantity, balance_before, balance_after, reference)
                    VALUES (?, ?, 'invoice', ?, ?, ?, ?)
                ''', (item['id'], datetime.now().strftime('%Y-%m-%d %H:%M:%S'), -item['qty'], old_qty, new_qty, f"Invoice {invoice_no}"))
                cursor.execute('''
                    INSERT INTO invoice_items (invoice_id, medicine_id, medicine_name, quantity, price_per_unit, total)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (invoice_id, item['id'], item['name'], item['qty'], item['price'], item['total']))
            
            conn.commit()
            conn.close()
            
            # Prepare PDF data
            pdf_data = {
                'invoice_no': invoice_no,
                'date': datetime.now().strftime('%Y-%m-%d %H:%M'),
                'customer_name': customer_data['name'],
                'customer_phone': customer_data['phone'],
                'customer_address': customer_data['address'],
                'due_date': customer_data['due_date'],
                'items': self.bill_items[:],
                'subtotal': subtotal,
                'tax_rate': tax_rate,
                'tax_amount': tax_amount,
                'total': total,
                'business_tin': customer_data.get('business_tin', ''),
                'customer_tin': customer_data.get('customer_tin', '')
            }
            
            # Always save PDF automatically
            self.save_invoice_pdf(pdf_data)
            
            # Ask to print
            if messagebox.askyesno("Invoice Created", 
                                f"Invoice {invoice_no} saved.\nTotal: {total:.2f} Birr\nDue: {customer_data['due_date']}\n\nPrint invoice?"):
                self.print_invoice_pdf(pdf_data)
            
            self.clear_bill()
            self.load_medicines()
            return True
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed: {str(e)}")
            return False
                
    def create_database(self):
        """Create SQLite database and tables if they don't exist"""
        try:
            db_path = self.get_database_path()
            
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            # Create medicines table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS medicines (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    category TEXT,
                    manufacturer TEXT,
                    batch_number TEXT,
                    expiry_date TEXT,
                    quantity INTEGER DEFAULT 0,
                    purchase_price REAL DEFAULT 0,
                    selling_price REAL DEFAULT 0,
                    min_stock INTEGER DEFAULT 10,
                    location TEXT,
                    date_added TEXT,
                    UNIQUE(name, batch_number)
                )
            ''')
            
            # Create sales table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS sales (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    medicine_id INTEGER,
                    medicine_name TEXT,
                    quantity INTEGER,
                    price_per_unit REAL,
                    total_price REAL,
                    sale_date TEXT,
                    customer_name TEXT,
                    FOREIGN KEY (medicine_id) REFERENCES medicines (id)
                )
            ''')
            try:
                cursor.execute("ALTER TABLE sales ADD COLUMN tin TEXT")
            except sqlite3.OperationalError:
                pass
            try:
                cursor.execute("ALTER TABLE sales ADD COLUMN mode TEXT")
            except sqlite3.OperationalError:
                pass
            try:
                cursor.execute("ALTER TABLE sales ADD COLUMN payment_method TEXT")
            except sqlite3.OperationalError:
                pass
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS stock_movements (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    medicine_id INTEGER NOT NULL,
                    date TEXT NOT NULL,
                    type TEXT NOT NULL,  -- 'purchase', 'sale', 'adjustment'
                    quantity INTEGER NOT NULL,
                    balance_before INTEGER NOT NULL,
                    balance_after INTEGER NOT NULL,
                    reference TEXT,      -- sale_id, purchase_id, etc.
                    notes TEXT,
                    FOREIGN KEY (medicine_id) REFERENCES medicines(id)
                )
            ''')
            
            # Invoices table (credit sales)
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS invoices (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    invoice_number TEXT UNIQUE NOT NULL,
                    customer_name TEXT,
                    customer_phone TEXT,
                    customer_address TEXT,
                    date_issued TEXT NOT NULL,
                    due_date TEXT NOT NULL,
                    subtotal REAL NOT NULL,
                    tax_rate REAL DEFAULT 15,
                    tax_amount REAL DEFAULT 0,
                    total REAL NOT NULL,
                    amount_paid REAL DEFAULT 0,
                    status TEXT DEFAULT 'Pending',  -- Pending, Partial, Paid, Overdue, Cancelled
                    notes TEXT,
                    created_by TEXT,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            # Invoice items (products sold in each invoice)
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS invoice_items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    invoice_id INTEGER NOT NULL,
                    medicine_id INTEGER NOT NULL,
                    medicine_name TEXT NOT NULL,
                    quantity INTEGER NOT NULL,
                    price_per_unit REAL NOT NULL,
                    total REAL NOT NULL,
                    FOREIGN KEY (invoice_id) REFERENCES invoices(id) ON DELETE CASCADE,
                    FOREIGN KEY (medicine_id) REFERENCES medicines(id)
                )
            ''')

            # Payments received for invoices
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS invoice_payments (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    invoice_id INTEGER NOT NULL,
                    amount REAL NOT NULL,
                    payment_date TEXT NOT NULL,
                    payment_method TEXT,
                    reference TEXT,
                    FOREIGN KEY (invoice_id) REFERENCES invoices(id) ON DELETE CASCADE
                )
            ''')
            
            conn.commit()
            conn.close()
            
            # Show welcome message (first time only)
            welcome_file = os.path.join(self.get_app_data_dir(), '.welcome_shown')
            if not os.path.exists(welcome_file):
                messagebox.showinfo("📍 Data Location", 
                                   f"Your pharmacy data will be stored at:\n\n{self.get_app_data_dir()}\n\n"
                                   f"This location is safe and won't be affected by app updates.\n"
                                   f"You can always find your data in File → Data Location")
                with open(welcome_file, 'w') as f:
                    f.write('welcome')
                    
        except Exception as e:
            messagebox.showerror("Database Error", f"Could not create database:\n{str(e)}")
    
    def setup_styles(self):
        """Configure styles"""
        style = ttk.Style()
        style.configure('Treeview.Heading', font=('Segoe UI', 10, 'bold'))
        style.configure('Treeview', font=('Segoe UI', 9), rowheight=25)
    
    def create_status_bar(self):
        """Create status bar - visible on all pages"""
        status_frame = tk.Frame(self.root, bg='#34495e', height=30)
        status_frame.pack(side='bottom', fill='x')
        status_frame.pack_propagate(False)
        
        # Left side - Date & Version
        left_text = f"📅 {datetime.now().strftime('%Y-%m-%d')}  |  v1.0"
        tk.Label(status_frame, text=left_text, bg='#34495e', fg='#bdc3c7',
                font=('Segoe UI', 9)).pack(side='left', padx=15, pady=5)
        
        # Center - Developer Info
        center_text = "👨‍💻 Merawi Yohannes  |  📱 0921-540-245  |  📱 0960-633-549"
        tk.Label(status_frame, text=center_text, bg='#34495e', fg='white',
                font=('Segoe UI', 9, 'bold')).pack(side='left', padx=20, pady=5)
        
        # Right side - Location & Support
        right_text = "📍 Addis Ababa, Ethiopia  |  ⚕️ Support 24/7"
        tk.Label(status_frame, text=right_text, bg='#34495e', fg='#bdc3c7',
                font=('Segoe UI', 9)).pack(side='right', padx=15, pady=5)
    
    def create_menu(self):
        """Create menu bar"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="📁 File", menu=file_menu)
        file_menu.add_command(label="💾 Backup to Excel", command=self.backup_to_excel)
        file_menu.add_command(label="📂 Import from Excel", command=self.import_from_excel)
        file_menu.add_separator()
        file_menu.add_command(label="📍 Data Location", command=self.show_data_location)
        file_menu.add_separator()
        file_menu.add_command(label="🚪 Exit", command=self.root.quit)
        
        # Reports menu
        reports_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="📊 Reports", menu=reports_menu)
        reports_menu.add_command(label="⚠️ Low Stock Report", command=self.low_stock_report)
        reports_menu.add_command(label="⏰ Expiring Medicines", command=self.expiring_report)
        reports_menu.add_command(label="💰 Today's Sales", command=self.today_sales_report)
        reports_menu.add_command(label="📊 Invoice Analysis", command=self.show_invoice_analysis)
        
        # Settings menu
        settings_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="⚙️ Settings", menu=settings_menu)
        settings_menu.add_command(label="Preferences", command=self.show_settings)
        
        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="❓ Help", menu=help_menu)
        help_menu.add_command(label="ℹ️ About", command=self.show_about)
    
    # ========== DASHBOARD TAB ==========
    
    def create_dashboard_tab(self):
        """Create dashboard with critical visual graphs - FULL WIDTH, ALL LABELS VISIBLE"""
        self.dashboard_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.dashboard_frame, text="📊 Dashboard")
        
        # ========== GET ALL DATA ==========
        conn = sqlite3.connect(self.get_database_path())
        cursor = conn.cursor()
        
        # Financial data
        cursor.execute('SELECT SUM(quantity * purchase_price) FROM medicines')
        total_cost = cursor.fetchone()[0] or 0
        
        cursor.execute('SELECT SUM(quantity * selling_price) FROM medicines')
        total_revenue_potential = cursor.fetchone()[0] or 0
        
        potential_profit = total_revenue_potential - total_cost
        profit_margin = (potential_profit / total_revenue_potential * 100) if total_revenue_potential > 0 else 0
        
        # Today's sales
        today = datetime.now().strftime('%Y-%m-%d')
        cursor.execute('SELECT SUM(total_price), COUNT(*) FROM sales WHERE sale_date LIKE ?', (f'{today}%',))
        today_data = cursor.fetchone()
        today_revenue = today_data[0] or 0
        today_transactions = today_data[1] or 0
        
        # All-time sales
        cursor.execute('SELECT SUM(total_price) FROM sales')
        total_revenue = cursor.fetchone()[0] or 0
        
        cursor.execute('SELECT COUNT(*) FROM sales')
        total_transactions = cursor.fetchone()[0] or 0
        
        # Counts
        cursor.execute('SELECT COUNT(*) FROM medicines')
        total_products = cursor.fetchone()[0]
        
        cursor.execute('SELECT SUM(quantity) FROM medicines')
        total_units = cursor.fetchone()[0] or 0
        
        # Low stock count
        cursor.execute('SELECT COUNT(*) FROM medicines WHERE quantity <= min_stock')
        low_stock_count = cursor.fetchone()[0]
        
        # ========== EXPIRY ANALYSIS ==========
        today_date = datetime.now().date()
        cursor.execute('SELECT expiry_date, quantity, min_stock, selling_price FROM medicines')
        all_meds = cursor.fetchall()
        
        # Status counts
        healthy = {'count': 0, 'units': 0, 'value': 0}
        low_stock = {'count': 0, 'units': 0, 'value': 0}
        expiring = {'count': 0, 'units': 0, 'value': 0}
        expired = {'count': 0, 'units': 0, 'value': 0, 'loss': 0}
        
        # Monthly expiry distribution (FULL YEAR)
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        monthly_expiry = {month: 0 for month in months}
        
        # Expiry value by month
        monthly_expiry_value = {month: 0 for month in months}
        
        for exp, qty, min_stk, price in all_meds:
            item_value = qty * price
            is_expired = False
            is_expiring = False
            
            if exp and len(exp) == 10:
                try:
                    expiry = datetime.strptime(exp, '%Y-%m-%d').date()
                    days = (expiry - today_date).days
                    month_name = expiry.strftime('%b')
                    
                    if month_name in monthly_expiry:
                        monthly_expiry[month_name] += qty
                        monthly_expiry_value[month_name] += item_value
                    
                    if days < 0:
                        expired['count'] += 1
                        expired['units'] += qty
                        expired['value'] += item_value
                        is_expired = True
                    elif days <= 30:
                        expiring['count'] += 1
                        expiring['units'] += qty
                        expiring['value'] += item_value
                        is_expiring = True
                except:
                    pass
            
            if not is_expired and not is_expiring:
                if qty <= min_stk:
                    low_stock['count'] += 1
                    low_stock['units'] += qty
                    low_stock['value'] += item_value
                else:
                    healthy['count'] += 1
                    healthy['units'] += qty
                    healthy['value'] += item_value
        
        # ========== CATEGORY DATA ==========
        cursor.execute('''
            SELECT category, COUNT(*) as count, SUM(quantity) as total_qty 
            FROM medicines 
            WHERE category IS NOT NULL AND category != ''
            GROUP BY category 
            ORDER BY count DESC 
        ''')
        categories = cursor.fetchall()
        
        # ========== TOP SELLERS ==========
        cursor.execute('''
            SELECT medicine_name, SUM(quantity) as total_sold, SUM(total_price) as revenue
            FROM sales 
            GROUP BY medicine_name 
            ORDER BY total_sold DESC 
            LIMIT 5
        ''')
        top_sellers = cursor.fetchall()
        
        # ========== MONTHLY SALES TREND (12 MONTHS) ==========
        monthly_sales = {}
        monthly_transactions = {}
        
        # Initialize last 12 months
        for i in range(11, -1, -1):
            month_date = today_date - timedelta(days=30*i)
            month_key = month_date.strftime('%b %Y')
            monthly_sales[month_key] = 0
            monthly_transactions[month_key] = 0
        
        cursor.execute('''
            SELECT strftime('%Y-%m', sale_date) as month, 
                SUM(total_price) as total,
                COUNT(*) as trans
            FROM sales 
            GROUP BY month 
            ORDER BY month DESC 
            LIMIT 12
        ''')
        for month, amount, trans in cursor.fetchall():
            month_name = datetime.strptime(month + '-01', '%Y-%m-%d').strftime('%b %Y')
            monthly_sales[month_name] = amount or 0
            monthly_transactions[month_name] = trans or 0
        
        # ========== TOP MANUFACTURERS ==========
        cursor.execute('''
            SELECT manufacturer, COUNT(*) as count 
            FROM medicines 
            WHERE manufacturer IS NOT NULL AND manufacturer != ''
            GROUP BY manufacturer 
            ORDER BY count DESC 
            LIMIT 5
        ''')
        top_manufacturers = cursor.fetchall()
        
        # ========== PRICE RANGE DISTRIBUTION ==========
        price_ranges = {'0-50': 0, '51-100': 0, '101-200': 0, '201-500': 0, '500+': 0}
        cursor.execute('SELECT selling_price, quantity FROM medicines')
        for price, qty in cursor.fetchall():
            if price <= 50:
                price_ranges['0-50'] += qty
            elif price <= 100:
                price_ranges['51-100'] += qty
            elif price <= 200:
                price_ranges['101-200'] += qty
            elif price <= 500:
                price_ranges['201-500'] += qty
            else:
                price_ranges['500+'] += qty
        
        # ========== WEEKDAY SALES PATTERN ==========
        weekdays = {'Mon': 0, 'Tue': 0, 'Wed': 0, 'Thu': 0, 'Fri': 0, 'Sat': 0, 'Sun': 0}
        cursor.execute('SELECT sale_date, total_price FROM sales')
        for date_str, amount in cursor.fetchall():
            if date_str and len(date_str) >= 10:
                try:
                    sale_date = datetime.strptime(date_str[:10], '%Y-%m-%d').date()
                    weekday = sale_date.strftime('%a')
                    if weekday in weekdays:
                        weekdays[weekday] += amount
                except:
                    pass
        
        conn.close()
        
        # ========== HEADER ==========
        header_frame = tk.Frame(self.dashboard_frame, bg='white', padx=25, pady=15)
        header_frame.pack(fill='x')
        
        title_frame = tk.Frame(header_frame, bg='white')
        title_frame.pack(side='left')
        
        tk.Label(title_frame, text="📊 MERawi PHARMACY PRO", 
                bg='white', fg='#2c3e50', font=('Segoe UI', 18, 'bold')).pack(anchor='w')
        tk.Label(title_frame, text="Business Intelligence Dashboard - Complete Pharmacy Analytics", 
                bg='white', fg='#7f8c8d', font=('Segoe UI', 10)).pack(anchor='w')
        
        refresh_btn = tk.Button(header_frame, text="🔄 Refresh Data", bg='#3498db', fg='white',
                            font=('Segoe UI', 10, 'bold'), padx=20, pady=8,
                            relief='flat', command=self.refresh_dashboard)
        refresh_btn.pack(side='right')
        
        tk.Label(header_frame, text=f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                bg='white', fg='#95a5a6', font=('Segoe UI', 9)).pack(anchor='w', pady=(5,0))
        
        # ========== SCROLLABLE AREA (FULL WIDTH) ==========
        canvas = tk.Canvas(self.dashboard_frame, bg='#f5f5f5', highlightthickness=0)
        scrollbar = tk.Scrollbar(self.dashboard_frame, orient='vertical', command=canvas.yview)
        scrollable = tk.Frame(canvas, bg='#f5f5f5')
        
        scrollable.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=scrollable, anchor='nw', width=canvas.winfo_width())
        
        def configure_width(e):
            canvas.itemconfig(1, width=e.width)
        canvas.bind('<Configure>', configure_width)
        
        canvas.configure(yscrollcommand=scrollbar.set)
        
        def wheel(e): canvas.yview_scroll(int(-1*(e.delta/120)), 'units')
        scrollable.bind('<Enter>', lambda e: canvas.bind_all('<MouseWheel>', wheel))
        scrollable.bind('<Leave>', lambda e: canvas.unbind_all('<MouseWheel>'))
        
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # ========== MAIN CONTENT (FULL WIDTH) ==========
        main = tk.Frame(scrollable, bg='#f5f5f5', padx=25, pady=25)
        main.pack(fill='both', expand=True)
        
        # ========== KPI CARDS ROW 1 ==========
        kpi_frame = tk.Frame(main, bg='#f5f5f5')
        kpi_frame.pack(fill='x', pady=5)
        
        cards = [
            ("💰 TOTAL INVESTMENT", f"{total_cost:,.0f} Birr", "What you paid", '#e74c3c'),
            ("📈 POTENTIAL REVENUE", f"{total_revenue_potential:,.0f} Birr", "Customer value", '#2ecc71'),
            ("💵 POTENTIAL PROFIT", f"{potential_profit:,.0f} Birr", f"Margin: {profit_margin:.1f}%", '#3498db'),
            ("📊 TODAY'S SALES", f"{today_revenue:,.0f} Birr", f"{today_transactions} transactions", '#f39c12')
        ]
        
        for title, value, sub, color in cards:
            card = tk.Frame(kpi_frame, bg='white', padx=15, pady=12, relief='ridge', bd=1)
            card.pack(side='left', expand=True, fill='both', padx=3)
            tk.Label(card, text=title, bg='white', fg='#7f8c8d', font=('Segoe UI', 9)).pack()
            tk.Label(card, text=value, bg='white', fg=color, font=('Segoe UI', 14, 'bold')).pack()
            tk.Label(card, text=sub, bg='white', fg='#95a5a6', font=('Segoe UI', 8)).pack()
        
        # ========== KPI CARDS ROW 2 ==========
        kpi2_frame = tk.Frame(main, bg='#f5f5f5')
        kpi2_frame.pack(fill='x', pady=5)
        
        cards2 = [
            ("💊 TOTAL PRODUCTS", f"{total_products}", f"{int(total_units):,} units", '#2c3e50'),
            ("⚠️ LOW STOCK", f"{low_stock['count']}", f"{low_stock['units']} units", '#e67e22'),
            ("⏰ EXPIRING SOON", f"{expiring['count']}", f"{expiring['units']} units (30 days)", '#e74c3c'),
            ("❌ EXPIRED", f"{expired['count']}", f"{expired['units']} units", '#c0392b'),
            ("✅ HEALTHY", f"{healthy['count']}", f"{healthy['units']} units", '#27ae60')
        ]
        
        for title, value, sub, color in cards2:
            card = tk.Frame(kpi2_frame, bg='white', padx=15, pady=12, relief='ridge', bd=1)
            card.pack(side='left', expand=True, fill='both', padx=3)
            tk.Label(card, text=title, bg='white', fg='#7f8c8d', font=('Segoe UI', 9)).pack()
            tk.Label(card, text=value, bg='white', fg=color, font=('Segoe UI', 14, 'bold')).pack()
            tk.Label(card, text=sub, bg='white', fg='#95a5a6', font=('Segoe UI', 8)).pack()
        
        # ========== GRAPH 1-2: FINANCIAL HEALTH + INVENTORY STATUS ==========
        graph_row1 = tk.Frame(main, bg='#f5f5f5')
        graph_row1.pack(fill='x', pady=10)
        
        # Graph 1: Financial Pie (LEFT)
        frame1 = tk.LabelFrame(graph_row1, text="💰 Financial Health - Where Your Money Is", 
                            bg='white', padx=20, pady=15, font=('Segoe UI', 11, 'bold'))
        frame1.pack(side='left', fill='both', expand=True, padx=3)
        
        fig1 = Figure(figsize=(6, 4), dpi=100)
        fig1.patch.set_facecolor('white')
        ax1 = fig1.add_subplot(111)
        
        if total_cost > 0 or potential_profit > 0:
            wedges, texts, autos = ax1.pie(
                [total_cost, potential_profit], 
                labels=[f'Cost: {total_cost:,.0f}', f'Profit: {potential_profit:,.0f}'], 
                colors=['#e74c3c', '#2ecc71'],
                autopct=lambda pct: f'{pct:.1f}%\n({pct*total_revenue_potential/100:,.0f})',
                textprops={'fontsize': 9}
            )
            for text in texts:
                text.set_fontsize(9)
            for auto in autos:
                auto.set_fontsize(8)
                auto.set_color('white')
                auto.set_fontweight('bold')
        else:
            ax1.text(0.5, 0.5, 'No financial data', ha='center', va='center')
        
        ax1.set_title('Investment Breakdown\n(What you paid vs What you gain)', fontsize=10, fontweight='bold')
        canvas1 = FigureCanvasTkAgg(fig1, frame1)
        canvas1.draw()
        canvas1.get_tk_widget().pack(fill='both', expand=True)
        
        # Graph 2: Inventory Status Pie (RIGHT)
        frame2 = tk.LabelFrame(graph_row1, text="📦 Inventory Health - Stock Status", 
                            bg='white', padx=20, pady=15, font=('Segoe UI', 11, 'bold'))
        frame2.pack(side='right', fill='both', expand=True, padx=3)
        
        fig2 = Figure(figsize=(6, 4), dpi=100)
        fig2.patch.set_facecolor('white')
        ax2 = fig2.add_subplot(111)
        
        if healthy['units'] + low_stock['units'] + expiring['units'] + expired['units'] > 0:
            data = [healthy['units'], low_stock['units'], expiring['units'], expired['units']]
            labels = [
                f'Healthy: {healthy["units"]}', 
                f'Low Stock: {low_stock["units"]}', 
                f'Expiring: {expiring["units"]}', 
                f'Expired: {expired["units"]}'
            ]
            colors = ['#2ecc71', '#f39c12', '#e67e22', '#e74c3c']
            
            wedges, texts, autos = ax2.pie(
                data, labels=labels, colors=colors,
                autopct=lambda pct: f'{pct:.1f}%',
                textprops={'fontsize': 9}
            )
            for text in texts:
                text.set_fontsize(9)
            for auto in autos:
                auto.set_fontsize(8)
                auto.set_color('white')
                auto.set_fontweight('bold')
        else:
            ax2.text(0.5, 0.5, 'No inventory data', ha='center', va='center')
        
        ax2.set_title('Units by Status\n(Good ⚡ Low ⏰ Expiring ❌ Expired)', fontsize=10, fontweight='bold')
        canvas2 = FigureCanvasTkAgg(fig2, frame2)
        canvas2.draw()
        canvas2.get_tk_widget().pack(fill='both', expand=True)
        
        # ========== GRAPH 3: EXPIRY WARNING BAR CHART (FULL WIDTH) ==========
        frame3 = tk.LabelFrame(main, text="⏰ Expiry Calendar - Units Expiring Each Month", 
                            bg='white', padx=20, pady=15, font=('Segoe UI', 11, 'bold'))
        frame3.pack(fill='x', pady=10)
        
        fig3 = Figure(figsize=(12, 4), dpi=100)
        fig3.patch.set_facecolor('white')
        ax3 = fig3.add_subplot(111)
        
        month_list = list(monthly_expiry.keys())
        values = list(monthly_expiry.values())
        value_labels = [f'{v:,}' for v in values]
        
        if sum(values) > 0:
            colors = ['#e74c3c' if i == datetime.now().month-1 else '#f39c12' if val > 0 else '#bdc3c7' 
                    for i, val in enumerate(values)]
            
            bars = ax3.bar(month_list, values, color=colors, width=0.8)
            ax3.set_title('UNITS EXPIRING EACH MONTH - PLAN YOUR SALES!', fontsize=12, fontweight='bold')
            ax3.set_ylabel('Number of Units', fontsize=10)
            ax3.set_xlabel('Month', fontsize=10)
            ax3.grid(axis='y', alpha=0.3)
            
            # Add value labels on top of bars
            for bar, val, label in zip(bars, values, value_labels):
                if val > 0:
                    ax3.text(bar.get_x() + bar.get_width()/2., bar.get_height() + max(values)*0.01,
                            label, ha='center', va='bottom', fontsize=9, fontweight='bold')
            
            # Highlight current month
            current_month = datetime.now().strftime('%b')
            ax3.text(0.02, 0.95, f'🔴 Current Month: {current_month} - ACT NOW!', 
                    transform=ax3.transAxes, fontsize=10, color='#e74c3c', fontweight='bold')
        else:
            ax3.text(0.5, 0.5, 'No expiry data available', ha='center', va='center')
        
        plt.setp(ax3.xaxis.get_majorticklabels(), rotation=0)
        fig3.tight_layout()
        
        canvas3 = FigureCanvasTkAgg(fig3, frame3)
        canvas3.draw()
        canvas3.get_tk_widget().pack(fill='both', expand=True)
        
        # ========== GRAPH 4-5: TOP SELLERS + CATEGORIES ==========
        graph_row2 = tk.Frame(main, bg='#f5f5f5')
        graph_row2.pack(fill='x', pady=10)
        
        # Graph 4: Top Sellers (LEFT)
        frame4 = tk.LabelFrame(graph_row2, text="🏆 Top 5 Best Selling Products", 
                            bg='white', padx=20, pady=15, font=('Segoe UI', 11, 'bold'))
        frame4.pack(side='left', fill='both', expand=True, padx=3)
        
        fig4 = Figure(figsize=(6, 4), dpi=100)
        fig4.patch.set_facecolor('white')
        ax4 = fig4.add_subplot(111)
        
        if top_sellers:
            products = [p[0][:20] + '..' if len(p[0]) > 20 else p[0] for p in top_sellers]
            quantities = [p[1] for p in top_sellers]
            revenues = [p[2] for p in top_sellers]
            
            bars = ax4.barh(products, quantities, color='#3498db')
            ax4.set_xlabel('Quantity Sold', fontsize=9)
            ax4.set_title('Most Popular Medicines', fontsize=10, fontweight='bold')
            
            for i, (bar, qty, rev) in enumerate(zip(bars, quantities, revenues)):
                ax4.text(qty + max(quantities)*0.01, bar.get_y() + bar.get_height()/2,
                        f'{qty} units\n({rev:,.0f} Birr)', va='center', fontsize=8)
        else:
            ax4.text(0.5, 0.5, 'No sales data yet', ha='center', va='center')
        
        fig4.tight_layout()
        canvas4 = FigureCanvasTkAgg(fig4, frame4)
        canvas4.draw()
        canvas4.get_tk_widget().pack(fill='both', expand=True)
        
        # Graph 5: Categories (RIGHT)
        frame5 = tk.LabelFrame(graph_row2, text="📊 Products by Category", 
                            bg='white', padx=20, pady=15, font=('Segoe UI', 11, 'bold'))
        frame5.pack(side='right', fill='both', expand=True, padx=3)
        
        fig5 = Figure(figsize=(6, 4), dpi=100)
        fig5.patch.set_facecolor('white')
        ax5 = fig5.add_subplot(111)
        
        if categories:
            cat_names = [c[0][:15] + '..' if len(c[0]) > 15 else c[0] for c in categories[:5]]
            cat_counts = [c[1] for c in categories[:5]]
            cat_units = [c[2] for c in categories[:5]]
            
            colors = plt.cm.Set3(np.linspace(0, 1, len(cat_names)))
            bars = ax5.bar(cat_names, cat_counts, color=colors)
            ax5.set_title('Inventory Distribution by Category', fontsize=10, fontweight='bold')
            ax5.set_ylabel('Number of Products', fontsize=9)
            
            for bar, count, units in zip(bars, cat_counts, cat_units):
                height = bar.get_height()
                ax5.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                        f'{count} products\n({units} units)', ha='center', va='bottom', fontsize=8)
        else:
            ax5.text(0.5, 0.5, 'No category data', ha='center', va='center')
        
        plt.setp(ax5.xaxis.get_majorticklabels(), rotation=15)
        fig5.tight_layout()
        canvas5 = FigureCanvasTkAgg(fig5, frame5)
        canvas5.draw()
        canvas5.get_tk_widget().pack(fill='both', expand=True)
        
        # ========== GRAPH 6-7: SALES TREND + WEEKDAY PATTERN ==========
        graph_row3 = tk.Frame(main, bg='#f5f5f5')
        graph_row3.pack(fill='x', pady=10)
        
        # Graph 6: Monthly Sales Trend (LEFT)
        frame6 = tk.LabelFrame(graph_row3, text="📈 Sales Performance - Last 12 Months", 
                            bg='white', padx=20, pady=15, font=('Segoe UI', 11, 'bold'))
        frame6.pack(side='left', fill='both', expand=True, padx=3)
        
        fig6 = Figure(figsize=(7, 4), dpi=100)
        fig6.patch.set_facecolor('white')
        ax6 = fig6.add_subplot(111)
        
        months_list = list(monthly_sales.keys())
        sales_values = list(monthly_sales.values())
        
        if sum(sales_values) > 0:
            ax6.plot(months_list, sales_values, marker='o', linewidth=2, markersize=8, color='#e67e22')
            ax6.fill_between(months_list, sales_values, alpha=0.3, color='#f39c12')
            ax6.set_title('Monthly Revenue Trend', fontsize=10, fontweight='bold')
            ax6.set_ylabel('Sales (Birr)', fontsize=9)
            ax6.grid(True, alpha=0.3)
            
            for i, (month, val) in enumerate(zip(months_list, sales_values)):
                if val > 0:
                    ax6.annotate(f'{val:,.0f}', (month, val), textcoords="offset points", 
                            xytext=(0,10), ha='center', fontsize=8, fontweight='bold')
            
            # Add trend indicator
            if len(sales_values) >= 2:
                trend = sales_values[-1] - sales_values[0]
                trend_text = f'📊 12-Month Trend: {"▲ +" if trend > 0 else "▼ "}{abs(trend):,.0f} Birr'
                trend_color = '#2ecc71' if trend > 0 else '#e74c3c'
                ax6.text(0.02, 0.95, trend_text, transform=ax6.transAxes, 
                        fontsize=9, color=trend_color, fontweight='bold')
        else:
            ax6.text(0.5, 0.5, 'No sales data', ha='center', va='center')
        
        plt.setp(ax6.xaxis.get_majorticklabels(), rotation=45)
        fig6.tight_layout()
        canvas6 = FigureCanvasTkAgg(fig6, frame6)
        canvas6.draw()
        canvas6.get_tk_widget().pack(fill='both', expand=True)
        
        # Graph 7: Weekday Sales Pattern (RIGHT)
        frame7 = tk.LabelFrame(graph_row3, text="📅 Best Selling Days", 
                            bg='white', padx=20, pady=15, font=('Segoe UI', 11, 'bold'))
        frame7.pack(side='right', fill='both', expand=True, padx=3)
        
        fig7 = Figure(figsize=(5, 4), dpi=100)
        fig7.patch.set_facecolor('white')
        ax7 = fig7.add_subplot(111)
        
        days = list(weekdays.keys())
        day_values = list(weekdays.values())
        
        if sum(day_values) > 0:
            colors = ['#3498db' if i < 5 else '#e67e22' for i in range(7)]
            bars = ax7.bar(days, day_values, color=colors)
            ax7.set_title('Sales by Day of Week', fontsize=10, fontweight='bold')
            ax7.set_ylabel('Sales (Birr)', fontsize=9)
            
            for bar, val in zip(bars, day_values):
                if val > 0:
                    ax7.text(bar.get_x() + bar.get_width()/2., bar.get_height() + max(day_values)*0.01,
                            f'{val:,.0f}', ha='center', va='bottom', fontsize=8, fontweight='bold')
            
            # Find best day
            best_day = days[day_values.index(max(day_values))]
            ax7.text(0.05, 0.95, f'⭐ Best Day: {best_day}', transform=ax7.transAxes, 
                    fontsize=9, color='#f39c12', fontweight='bold')
        else:
            ax7.text(0.5, 0.5, 'No sales data', ha='center', va='center')
        
        fig7.tight_layout()
        canvas7 = FigureCanvasTkAgg(fig7, frame7)
        canvas7.draw()
        canvas7.get_tk_widget().pack(fill='both', expand=True)
        
        # ========== GRAPH 8: PRICE RANGE DISTRIBUTION ==========
        frame8 = tk.LabelFrame(main, text="💰 Price Range Distribution - Where Your Stock Value Is", 
                            bg='white', padx=20, pady=15, font=('Segoe UI', 11, 'bold'))
        frame8.pack(fill='x', pady=10)
        
        fig8 = Figure(figsize=(12, 3.5), dpi=100)
        fig8.patch.set_facecolor('white')
        ax8 = fig8.add_subplot(111)
        
        ranges = list(price_ranges.keys())
        range_values = list(price_ranges.values())
        
        if sum(range_values) > 0:
            colors = plt.cm.viridis(np.linspace(0, 1, len(ranges)))
            bars = ax8.bar(ranges, range_values, color=colors, width=0.7)
            ax8.set_title('Stock Distribution by Price Range', fontsize=11, fontweight='bold')
            ax8.set_ylabel('Number of Units', fontsize=9)
            ax8.set_xlabel('Price Range (Birr)', fontsize=9)
            
            for bar, val in zip(bars, range_values):
                if val > 0:
                    ax8.text(bar.get_x() + bar.get_width()/2., bar.get_height() + max(range_values)*0.01,
                            f'{val:,} units', ha='center', va='bottom', fontsize=9, fontweight='bold')
        else:
            ax8.text(0.5, 0.5, 'No price data', ha='center', va='center')
        
        fig8.tight_layout()
        canvas8 = FigureCanvasTkAgg(fig8, frame8)
        canvas8.draw()
        canvas8.get_tk_widget().pack(fill='both', expand=True)
        
        # ========== LOSS WARNING SECTION ==========
        if expired['units'] > 0 or expiring['units'] > 0:
            warn_frame = tk.Frame(main, bg='#fdeded', padx=20, pady=15)
            warn_frame.pack(fill='x', pady=10)
            
            tk.Label(warn_frame, text="⚠️ CRITICAL WARNING - FINANCIAL LOSS DETECTED", 
                    bg='#fdeded', fg='#c0392b', font=('Segoe UI', 13, 'bold')).pack()
            
            if expired['units'] > 0:
                loss_text = f"❌ You have already LOST {expired['value']:,.0f} Birr from {expired['units']} expired units!"
                tk.Label(warn_frame, text=loss_text, bg='#fdeded', fg='#c0392b',
                        font=('Segoe UI', 11)).pack()
            
            if expiring['units'] > 0:
                risk_text = f"⏰ {expiring['units']} units worth {expiring['value']:,.0f} Birr will expire in 30 days - SELL NOW!"
                tk.Label(warn_frame, text=risk_text, bg='#fdeded', fg='#e67e22',
                        font=('Segoe UI', 11, 'bold')).pack()
        
        # ========== ACTION ITEMS ==========
        action_frame = tk.Frame(main, bg='#e8f4f8', padx=20, pady=12)
        action_frame.pack(fill='x', pady=10)
        
        tk.Label(action_frame, text="🎯 YOUR NEXT ACTIONS:", bg='#e8f4f8', fg='#2980b9',
                font=('Segoe UI', 12, 'bold')).pack(anchor='w')
        
        actions = []
        if low_stock['units'] > 0:
            actions.append(f"📦 Reorder {low_stock['count']} low stock items ({low_stock['units']} units)")
        if expiring['units'] > 0:
            actions.append(f"🏷️ RUN SALE on {expiring['count']} products expiring soon")
        if expired['units'] > 0:
            actions.append(f"🗑️ Remove {expired['count']} expired products immediately")
        if len(top_sellers) > 0:
            actions.append(f"⭐ Stock more of your top sellers")
        
        if actions:
            for action in actions:
                tk.Label(action_frame, text=action, bg='#e8f4f8', fg='#2c3e50',
                        font=('Segoe UI', 10), anchor='w').pack(anchor='w', pady=2)
        else:
            tk.Label(action_frame, text="✅ All systems optimal! Keep up the good work!",
                    bg='#e8f4f8', fg='#27ae60', font=('Segoe UI', 11, 'bold')).pack()
        
        # ========== FOOTER ==========
        footer = tk.Frame(main, bg='#f5f5f5')
        footer.pack(fill='x', pady=(15, 5))
        
        tk.Label(footer, text="MERawi Pharmacy Pro - Business Intelligence Dashboard", 
                bg='#f5f5f5', fg='#7f8c8d', font=('Segoe UI', 9)).pack()
        tk.Label(footer, text="Merawi Yohannes · 0921-540-245 · merawiyohannes@gmail.com", 
                bg='#f5f5f5', fg='#7f8c8d', font=('Segoe UI', 9)).pack()
    
    def create_bin_card_tab(self):
        self.bin_card_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.bin_card_frame, text="📦 Bin Card")

        # ========== LEFT PANEL: Medicine list with search ==========
        left = tk.Frame(self.bin_card_frame, bg='white')
        left.pack(side='left', fill='y', padx=5, pady=5)

        tk.Label(left, text="Search Medicine", font=('Segoe UI', 11, 'bold')).pack(pady=(5,0))
        self.bin_search_var = tk.StringVar()
        bin_search_entry = tk.Entry(left, textvariable=self.bin_search_var, width=30)
        bin_search_entry.pack(pady=5)
        bin_search_entry.bind('<KeyRelease>', self.on_bin_search)

        self.bin_medicine_listbox = tk.Listbox(left, width=35, height=20)
        self.bin_medicine_listbox.pack(fill='both', expand=True, padx=5, pady=5)
        self.bin_medicine_listbox.bind('<<ListboxSelect>>', self.on_bin_medicine_select)

        # Add a Refresh button to reload the medicine list (in case new medicines were added)
        refresh_btn = tk.Button(left, text="🔄 Refresh List", bg='#3498db', fg='white',
                                font=('Segoe UI', 9), command=self.load_all_medicines)
        refresh_btn.pack(pady=5)

        # ========== RIGHT PANEL: Movement history with scrollbar ==========
        right = tk.Frame(self.bin_card_frame, bg='white')
        right.pack(side='right', fill='both', expand=True, padx=5, pady=5)

        tk.Label(right, text="Movement History", font=('Segoe UI', 11, 'bold')).pack(pady=5)

        # Create a frame to hold the Treeview and scrollbar
        tree_frame = tk.Frame(right)
        tree_frame.pack(fill='both', expand=True)

        columns = ('Date', 'Type', 'Quantity', 'Balance', 'Reference')
        self.movement_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=20)
        for col in columns:
            self.movement_tree.heading(col, text=col)
            self.movement_tree.column(col, width=120)

        # Add vertical scrollbar
        v_scroll = ttk.Scrollbar(tree_frame, orient='vertical', command=self.movement_tree.yview)
        self.movement_tree.configure(yscrollcommand=v_scroll.set)

        # Pack tree and scrollbar
        self.movement_tree.pack(side='left', fill='both', expand=True)
        v_scroll.pack(side='right', fill='y')

        # Optional horizontal scrollbar (if needed)
        h_scroll = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.movement_tree.xview)
        self.movement_tree.configure(xscrollcommand=h_scroll.set)
        h_scroll.pack(side='bottom', fill='x')

        # Store all medicines for filtering
        self.all_medicines = []  # list of (id, name)
        self.load_all_medicines()
        self.update_bin_listbox()
    
    def load_all_medicines(self):
        """Load all medicine IDs and names into memory."""
        conn = sqlite3.connect(self.get_database_path())
        cursor = conn.cursor()
        cursor.execute('SELECT id, name FROM medicines ORDER BY name')
        self.all_medicines = cursor.fetchall()
        conn.close()

    def on_bin_search(self, event):
        """Filter the medicine list based on search text."""
        self.update_bin_listbox()

    def update_bin_listbox(self):
        """Update the listbox with medicines matching the search term."""
        search_term = self.bin_search_var.get().lower()
        self.bin_medicine_listbox.delete(0, tk.END)
        for mid, name in self.all_medicines:
            if search_term in name.lower():
                self.bin_medicine_listbox.insert(tk.END, f"{mid}|{name}")

    def on_bin_medicine_select(self, event):
        selection = self.bin_medicine_listbox.curselection()
        if not selection:
            return
        item = self.bin_medicine_listbox.get(selection[0])
        medicine_id = item.split('|')[0]
        # Clear tree
        for row in self.movement_tree.get_children():
            self.movement_tree.delete(row)
        # Fetch movements for this medicine
        conn = sqlite3.connect(self.get_database_path())
        cursor = conn.cursor()
        cursor.execute('''
            SELECT date, type, quantity, balance_after, reference
            FROM stock_movements
            WHERE medicine_id = ?
            ORDER BY date ASC
        ''', (medicine_id,))
        movements = cursor.fetchall()
        for row in movements:
            date, typ, qty, bal, ref = row
            display_qty = f"{qty:+d}"  # shows + or -
            self.movement_tree.insert('', 'end', values=(date, typ, display_qty, bal, ref))
        conn.close()
    
    # ========== MEDICINES TAB ==========
    
    def check_owner_password(self, feature_name="this feature"):
        """Ask for owner password before allowing sensitive features."""
        dialog = tk.Toplevel(self.root)
        dialog.title("🔐 Owner Verification")
        dialog.geometry("300x150")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (150)
        y = (dialog.winfo_screenheight() // 2) - (75)
        dialog.geometry(f'+{x}+{y}')
        
        tk.Label(dialog, text=f"Enter password to access {feature_name}:",
                font=('Segoe UI', 10)).pack(pady=10)
        
        entry = tk.Entry(dialog, show="*", font=('Segoe UI', 11), width=20)
        entry.pack(pady=5)
        entry.focus_set()
        
        result = tk.BooleanVar(value=False)
        
        def verify():
            # Change this password to your own (owner sets it)
            if entry.get() == "merawi2024":   # ← CHANGE THIS PASSWORD!
                result.set(True)
                dialog.destroy()
            else:
                messagebox.showerror("Access Denied", "Incorrect password.\nAccess denied.")
                entry.delete(0, tk.END)
        
        tk.Button(dialog, text="OK", bg='#2c3e50', fg='white',
                font=('Segoe UI', 10), command=verify).pack(pady=5)
        dialog.bind('<Return>', lambda e: verify())
        
        self.root.wait_window(dialog)
        return result.get()
    
    def on_tab_changed(self, event):
        if getattr(self, '_switching_tab', False):
            return

        current = self.notebook.nametowidget(self.notebook.select())

        if current == self.dashboard_frame:
            self._switching_tab = True
            self.notebook.select(self._last_tab)
            self._switching_tab = False

            # 2. Ask for password
            if self.check_owner_password("Dashboard"):
                self._switching_tab = True
                self.notebook.select(self.dashboard_frame)
                self._switching_tab = False
                self._last_tab = self.dashboard_frame
            # else, stay on the previous tab (already done)
        else:
            self._last_tab = current
        
    def test_database_with_password(self):
        """Wrapper to check password before running test database."""
        if self.check_owner_password("Database Analytics"):
            self.test_database_contents()
    
    def create_medicines_tab(self):
        """Create medicines management tab"""
        self.medicines_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.medicines_frame, text="💊 Medicines")
        
        # Toolbar
        toolbar = tk.Frame(self.medicines_frame, bg='#ecf0f1', height=50)
        toolbar.pack(fill='x')
        toolbar.pack_propagate(False)
        
        tk.Button(toolbar, text="➕ Add Medicine", bg='#27ae60', fg='white',
                 font=('Segoe UI', 10), padx=10, command=self.add_medicine_dialog).pack(side='left', padx=5, pady=10)
        tk.Button(toolbar, text="✏️ Edit", bg='#2980b9', fg='white',
                 font=('Segoe UI', 10), padx=10, command=self.edit_medicine_dialog).pack(side='left', padx=5, pady=10)
        tk.Button(toolbar, text="🗑️ Delete", bg='#c0392b', fg='white',
                 font=('Segoe UI', 10), padx=10, command=self.delete_medicine).pack(side='left', padx=5, pady=10)
        tk.Button(toolbar, text="🔄 Refresh", bg='#f39c12', fg='white',
                 font=('Segoe UI', 10), padx=10, command=self.load_medicines).pack(side='left', padx=5, pady=10)
        tk.Button(toolbar, text="🔍 Test DB", bg='#9b59b6', fg='white',
                 font=('Segoe UI', 10), padx=10, command=self.test_database_with_password).pack(side='left', padx=5, pady=10)
        
        # Search
        search_frame = tk.Frame(self.medicines_frame, bg='white', relief='groove', bd=2)
        search_frame.pack(fill='x', padx=5, pady=5)
        
        tk.Label(search_frame, text="🔍 Search:", bg='white', font=('Segoe UI', 10)).pack(side='left', padx=5)
        self.search_entry = tk.Entry(search_frame, width=40, font=('Segoe UI', 10))
        self.search_entry.pack(side='left', padx=5, pady=5)
        self.search_entry.bind('<KeyRelease>', self.search_medicines)
        
        tk.Button(search_frame, text="Clear", bg='#95a5a6', fg='white',
                 font=('Segoe UI', 9), command=self.clear_search).pack(side='left', padx=5)
        
        # Treeview
        columns = ('ID', 'Name', 'Category', 'Batch', 'Expiry', 'Qty', 'Price', 'Location', 'Status')
        self.medicines_tree = ttk.Treeview(self.medicines_frame, columns=columns, show='headings', height=18)
        
        widths = [50, 200, 100, 100, 100, 70, 100, 100, 100]
        for i, col in enumerate(columns):
            self.medicines_tree.heading(col, text=col)
            self.medicines_tree.column(col, width=widths[i])
        
        scrollbar = ttk.Scrollbar(self.medicines_frame, orient='vertical', command=self.medicines_tree.yview)
        self.medicines_tree.configure(yscrollcommand=scrollbar.set)
        
        self.medicines_tree.pack(side='left', fill='both', expand=True, padx=5, pady=5)
        scrollbar.pack(side='right', fill='y')
        
        # Tags
        self.medicines_tree.tag_configure('expired', background='#ffcdd2')
        self.medicines_tree.tag_configure('expiring', background='#fff9c4')
        self.medicines_tree.tag_configure('lowstock', background='#ffe0b2')
    
    def load_medicines(self):
        """Load medicines into treeview"""
        for item in self.medicines_tree.get_children():
            self.medicines_tree.delete(item)
        
        try:
            conn = sqlite3.connect(self.get_database_path())
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM medicines ORDER BY name')
            medicines = cursor.fetchall()
            conn.close()
        except:
            medicines = []
        
        today = datetime.now().date()
        
        for med in medicines:
            status = "Good"
            tag = ''
            
            if med[5] and len(med[5]) == 10:
                try:
                    expiry = datetime.strptime(med[5], '%Y-%m-%d').date()
                    if expiry < today:
                        status = "EXPIRED"
                        tag = 'expired'
                    elif (expiry - today).days <= 30:
                        status = "Expiring Soon"
                        tag = 'expiring'
                except:
                    pass
            
            if status == "Good" and med[6] <= med[9]:
                status = "Low Stock"
                tag = 'lowstock'
            
            self.medicines_tree.insert('', 'end', values=(
                med[0], med[1], med[2] or "", med[4] or "", med[5] or "",
                med[6], f"{med[8]:.2f}", med[10] or "", status
            ), tags=(tag,))
    
    def search_medicines(self, event=None):
        """Search medicines"""
        search = self.search_entry.get().lower()
        
        for item in self.medicines_tree.get_children():
            self.medicines_tree.delete(item)
        
        try:
            conn = sqlite3.connect(self.get_database_path())
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM medicines WHERE LOWER(name) LIKE ? OR LOWER(category) LIKE ?',
                          (f'%{search}%', f'%{search}%'))
            medicines = cursor.fetchall()
            conn.close()
        except:
            medicines = []
        
        today = datetime.now().date()
        
        for med in medicines:
            status = "Good"
            tag = ''
            
            if med[5] and len(med[5]) == 10:
                try:
                    expiry = datetime.strptime(med[5], '%Y-%m-%d').date()
                    if expiry < today:
                        status = "EXPIRED"
                        tag = 'expired'
                    elif (expiry - today).days <= 30:
                        status = "Expiring Soon"
                        tag = 'expiring'
                except:
                    pass
            
            if status == "Good" and med[6] <= med[9]:
                status = "Low Stock"
                tag = 'lowstock'
            
            self.medicines_tree.insert('', 'end', values=(
                med[0], med[1], med[2] or "", med[4] or "", med[5] or "",
                med[6], f"{med[8]:.2f}", med[10] or "", status
            ), tags=(tag,))
    
    def clear_search(self):
        """Clear search"""
        self.search_entry.delete(0, tk.END)
        self.load_medicines()
    
    # ========== SALES TAB ==========
    
    def create_sales_tab(self):
        self.sales_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.sales_frame, text="💰 Sales")

        # Load configuration
        config = self.load_config()
        self.government_mode = config.get('government_mode', False)
        self.business_tin = config.get('tin', '')

        # ========== TOP TOOLBAR ==========
        top_bar = tk.Frame(self.sales_frame, bg='#34495e', height=40)
        top_bar.pack(fill='x')
        top_bar.pack_propagate(False)

        # Tax rate (only visible in private mode; in government mode it's fixed to 15%)
        if not self.government_mode:
            tk.Label(top_bar, text="Tax Rate:", bg='#34495e', fg='white',
                    font=('Segoe UI', 10)).pack(side='left', padx=20, pady=8)
            self.tax_var = tk.StringVar(value="15")
            tax_combo = ttk.Combobox(top_bar, textvariable=self.tax_var,
                                    values=["0", "5", "10", "15"], width=5)
            tax_combo.pack(side='left', padx=5)
            tk.Label(top_bar, text="%", bg='#34495e', fg='white').pack(side='left')
        else:
            # Fixed VAT at 15% (store the value for calculations)
            self.tax_var = tk.StringVar(value="15")
            # Optionally display a label
            tk.Label(top_bar, text="VAT 15% (Fixed)", bg='#34495e', fg='white',
                    font=('Segoe UI', 10)).pack(side='left', padx=20, pady=8)

        # Government‑specific fields: Customer TIN (optional) and Payment Method
        if self.government_mode:
            # Customer TIN
            tin_frame = tk.Frame(top_bar, bg='#34495e')
            tin_frame.pack(side='left', padx=20)
            tk.Label(tin_frame, text="Customer TIN:", bg='#34495e', fg='white',
                    font=('Segoe UI', 9)).pack(side='left')
            self.customer_tin = tk.Entry(tin_frame, width=15, bg='white')
            self.customer_tin.pack(side='left', padx=5)

            # Payment Method
            payment_frame = tk.Frame(top_bar, bg='#34495e')
            payment_frame.pack(side='left', padx=20)
            tk.Label(payment_frame, text="Payment:", bg='#34495e', fg='white',
                    font=('Segoe UI', 9)).pack(side='left')
            self.payment_method = ttk.Combobox(payment_frame, values=["Cash", "Card", "Transfer", "Other"],
                                            width=8, state='readonly')
            self.payment_method.set("Cash")
            self.payment_method.pack(side='left', padx=5)
        else:
            self.customer_tin = None
            self.payment_method = None

        # ========== MAIN PANED WINDOW (unchanged) ==========
        paned = ttk.PanedWindow(self.sales_frame, orient='horizontal')
        paned.pack(fill='both', expand=True, padx=5, pady=5)

        # LEFT PANEL – product search (same for both modes)
        left = tk.Frame(paned, bg='white', relief='groove', bd=1)
        paned.add(left, weight=1)

        tk.Label(left, text="🔍 Search Products", font=('Segoe UI', 11, 'bold'),
                bg='white', fg='#2c3e50').pack(pady=10)

        self.sale_search_entry = tk.Entry(left, width=30, font=('Segoe UI', 11))
        self.sale_search_entry.pack(pady=5)
        self.sale_search_entry.focus_set()
        self.sale_search_entry.bind('<KeyRelease>', self.search_sale_medicines)

        self.search_listbox = tk.Listbox(left, height=8, font=('Segoe UI', 10))
        self.search_listbox.pack(fill='both', expand=True, padx=10, pady=10)
        self.search_listbox.bind('<<ListboxSelect>>', self.on_medicine_select)

        qty_frame = tk.Frame(left, bg='white')
        qty_frame.pack(pady=10)

        tk.Label(qty_frame, text="Quantity:", font=('Segoe UI', 10)).pack(side='left', padx=5)
        self.sale_qty = tk.Entry(qty_frame, width=10, font=('Segoe UI', 11))
        self.sale_qty.pack(side='left', padx=5)

        tk.Button(qty_frame, text="Add to Bill", bg='#27ae60', fg='white',
                font=('Segoe UI', 10), padx=15, command=self.add_to_bill).pack(side='left', padx=5)

        # RIGHT PANEL – bill area
        right = tk.Frame(paned, bg='white', relief='groove', bd=1)
        paned.add(right, weight=1)

        tk.Label(right, text="CURRENT BILL", font=('Segoe UI', 11, 'bold'),
                bg='white', fg='#2c3e50').pack(pady=10)

        columns = ('Item', 'Product', 'Price', 'Qty', 'Total')
        self.bill_tree = ttk.Treeview(right, columns=columns, show='headings', height=12)

        widths = [50, 200, 80, 50, 100]
        for i, col in enumerate(columns):
            self.bill_tree.heading(col, text=col)
            self.bill_tree.column(col, width=widths[i])

        self.bill_tree.pack(fill='both', expand=True, padx=5, pady=5)

        # SUMMARY AREA
        summary = tk.Frame(right, bg='#f8f9fa', padx=15, pady=15)
        summary.pack(fill='x', padx=5, pady=5)

        subtotal_frame = tk.Frame(summary, bg='#f8f9fa')
        subtotal_frame.pack(fill='x', pady=2)
        tk.Label(subtotal_frame, text="Subtotal:", bg='#f8f9fa').pack(side='left')
        self.subtotal_label = tk.Label(subtotal_frame, text="0.00", bg='#f8f9fa', font=('Segoe UI', 10, 'bold'))
        self.subtotal_label.pack(side='right')

        tax_frame = tk.Frame(summary, bg='#f8f9fa')
        tax_frame.pack(fill='x', pady=2)
        tk.Label(tax_frame, text="VAT (15%):", bg='#f8f9fa').pack(side='left')
        self.tax_label = tk.Label(tax_frame, text="0.00", bg='#f8f9fa', font=('Segoe UI', 10, 'bold'))
        self.tax_label.pack(side='right')

        tk.Frame(summary, bg='#bdc3c7', height=1).pack(fill='x', pady=5)

        total_frame = tk.Frame(summary, bg='#f8f9fa')
        total_frame.pack(fill='x', pady=2)
        tk.Label(total_frame, text="TOTAL:", bg='#f8f9fa', font=('Segoe UI', 12, 'bold')).pack(side='left')
        self.total_label = tk.Label(total_frame, text="0.00 Birr", bg='#f8f9fa',
                                    font=('Segoe UI', 14, 'bold'), fg='#27ae60')
        self.total_label.pack(side='right')

        # BUTTONS
        btn_frame = tk.Frame(right, bg='white')
        btn_frame.pack(fill='x', padx=5, pady=5)

        tk.Button(btn_frame, text="❌ Remove", bg='#e67e22', fg='white',
                font=('Segoe UI', 9), command=self.remove_from_bill).pack(side='left', padx=2)
        tk.Button(btn_frame, text="🔄 Clear", bg='#7f8c8d', fg='white',
                font=('Segoe UI', 9), command=self.clear_bill).pack(side='left', padx=2)

        # The "Complete Sale" button text can also vary
        complete_text = "✅ Complete Sale" if not self.government_mode else "✅ Issue Fiscal Receipt"
        tk.Button(btn_frame, text=complete_text, bg='#27ae60', fg='white',
                font=('Segoe UI', 10, 'bold'), padx=15, command=self.complete_sale).pack(side='right', padx=2)
        tk.Button(btn_frame, text="📄 Credit Invoice", bg='#3498db', fg='white',
          font=('Segoe UI', 9, 'bold'), padx=10, command=self.create_invoice).pack(side='right', padx=2)
        self.bill_items = []
    
    def search_sale_medicines(self, event=None):
        """Search medicines for sale"""
        search = self.sale_search_entry.get().lower()
        self.search_listbox.delete(0, tk.END)
        
        if len(search) < 2:
            return
        
        try:
            conn = sqlite3.connect(self.get_database_path())
            cursor = conn.cursor()
            cursor.execute('''
                SELECT id, name, selling_price, quantity, expiry_date
                FROM medicines
                WHERE LOWER(name) LIKE ? AND quantity > 0
                ORDER BY name
            ''', (f'%{search}%',))
            medicines = cursor.fetchall()
            conn.close()
        except:
            medicines = []
        
        today = datetime.now().date()
        
        for med in medicines:
            if med[4] and len(med[4]) == 10:
                try:
                    expiry = datetime.strptime(med[4], '%Y-%m-%d').date()
                    if expiry < today:
                        continue
                except:
                    pass
            
            self.search_listbox.insert(tk.END, f"{med[0]}|{med[1]}|{med[2]}")
    
    def on_medicine_select(self, event=None):
        """Handle medicine selection"""
        if self.search_listbox.curselection():
            self.sale_qty.focus()
    
    def add_to_bill(self):
        """Add to bill"""
        if not self.search_listbox.curselection():
            messagebox.showwarning("Warning", "Please select a medicine")
            return
        
        try:
            qty = int(self.sale_qty.get())
            if qty <= 0:
                messagebox.showerror("Error", "Quantity must be positive")
                return
        except:
            messagebox.showerror("Error", "Invalid quantity")
            return
        
        med_info = self.search_listbox.get(self.search_listbox.curselection()[0])
        med_id, name, price = med_info.split('|')
        med_id = int(med_id)
        price = float(price)
        
        try:
            conn = sqlite3.connect(self.get_database_path())
            cursor = conn.cursor()
            cursor.execute('SELECT quantity FROM medicines WHERE id = ?', (med_id,))
            available = cursor.fetchone()[0]
            conn.close()
        except:
            available = 0
        
        if qty > available:
            messagebox.showerror("Error", f"Only {available} available")
            return
        
        total = price * qty
        
        self.bill_items.append({
            'id': med_id,
            'name': name,
            'price': price,
            'qty': qty,
            'total': total
        })
        
        self.bill_tree.insert('', 'end', values=(len(self.bill_items), name, f"{price:.2f}", qty, f"{total:.2f}"))
        self.update_bill_total()
        
        self.sale_search_entry.delete(0, tk.END)
        self.sale_qty.delete(0, tk.END)
        self.search_listbox.delete(0, tk.END)
    
    def remove_from_bill(self):
        """Remove from bill"""
        if not self.bill_tree.selection():
            messagebox.showwarning("Warning", "Please select an item")
            return
        
        selected = self.bill_tree.selection()[0]
        item_id = int(self.bill_tree.item(selected)['values'][0]) - 1
        
        if 0 <= item_id < len(self.bill_items):
            self.bill_items.pop(item_id)
            self.refresh_bill_display()
    
    def refresh_bill_display(self):
        """Refresh bill display"""
        for item in self.bill_tree.get_children():
            self.bill_tree.delete(item)
        
        for i, item in enumerate(self.bill_items, 1):
            self.bill_tree.insert('', 'end', values=(
                i, item['name'], f"{item['price']:.2f}", item['qty'], f"{item['total']:.2f}"
            ))
        
        self.update_bill_total()
    
    def update_bill_total(self):
        """Update bill total"""
        subtotal = sum(item['total'] for item in self.bill_items)
        tax_rate = float(self.tax_var.get()) / 100
        tax = subtotal * tax_rate
        total = subtotal + tax
        
        self.subtotal_label.config(text=f"{subtotal:.2f}")
        self.tax_label.config(text=f"{tax:.2f}")
        self.total_label.config(text=f"{total:.2f} Birr")
    
    def clear_bill(self):
        """Clear bill"""
        self.bill_items = []
        for item in self.bill_tree.get_children():
            self.bill_tree.delete(item)
        self.update_bill_total()
    
    def complete_sale(self):
        """Complete sale and record stock movements"""
        if not self.bill_items:
            messagebox.showwarning("Warning", "No items in bill")
            return
        
        if not messagebox.askyesno("Confirm", "Complete this sale?"):
            return
        
        try:
            conn = sqlite3.connect(self.get_database_path())
            cursor = conn.cursor()
            
            sale_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            receipt_no = f"R{datetime.now().strftime('%Y%m%d%H%M%S')}"
            mode = 'government' if self.government_mode else 'private'
            
            # Prepare fields that depend on mode
            customer_tin = ''
            payment_method = ''
            if self.government_mode:
                if self.customer_tin:
                    customer_tin = self.customer_tin.get().strip()
                if self.payment_method:
                    payment_method = self.payment_method.get()
            
            for item in self.bill_items:
                # Get current quantity before update
                cursor.execute('SELECT quantity FROM medicines WHERE id = ?', (item['id'],))
                old_qty = cursor.fetchone()[0]
                new_qty = old_qty - item['qty']
                
                # Update stock
                cursor.execute('UPDATE medicines SET quantity = ? WHERE id = ?', (new_qty, item['id']))
                
                # Insert stock movement
                cursor.execute('''
                    INSERT INTO stock_movements (medicine_id, date, type, quantity, balance_before, balance_after, reference)
                    VALUES (?, ?, 'sale', ?, ?, ?, ?)
                ''', (item['id'], sale_date, -item['qty'], old_qty, new_qty, f"Sale {receipt_no}"))
                
                # Insert sales record
                cursor.execute('''
                    INSERT INTO sales (medicine_id, medicine_name, quantity, price_per_unit, total_price, sale_date, tin, mode, payment_method)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (item['id'], item['name'], item['qty'], item['price'], item['total'], sale_date, customer_tin, mode, payment_method))
            
            conn.commit()
            conn.close()
            
            self.show_receipt()
            self.clear_bill()
            self.load_medicines()
            
        except Exception as e:
            messagebox.showerror("Error", f"Sale failed: {str(e)}")
            
    def show_receipt(self):
        receipt = tk.Toplevel(self.root)
        receipt.title("🧾 Receipt")
        receipt.geometry("400x600")
        receipt.configure(bg='white')

        # Header
        tk.Label(receipt, text="MERAWI PHARMACY PRO", bg='#2c3e50', fg='white',
                font=('Segoe UI', 14, 'bold'), pady=10).pack(fill='x')

        # Receipt info frame
        info_frame = tk.Frame(receipt, bg='white', padx=20, pady=10)
        info_frame.pack(fill='x')

        receipt_no = f"R{datetime.now().strftime('%Y%m%d%H%M%S')}"
        tk.Label(info_frame, text=f"Receipt No: {receipt_no}",
                bg='white', font=('Courier', 10)).pack(anchor='w')
        tk.Label(info_frame, text=f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                bg='white', font=('Courier', 10)).pack(anchor='w')

        if self.government_mode:
            # Show TINs and payment method
            tk.Label(info_frame, text=f"Business TIN: {self.business_tin}",
                    bg='white', font=('Courier', 10)).pack(anchor='w')
            customer_tin = self.customer_tin.get().strip() if self.customer_tin else ''
            if customer_tin:
                tk.Label(info_frame, text=f"Customer TIN: {customer_tin}",
                        bg='white', font=('Courier', 10)).pack(anchor='w')
            payment_method = self.payment_method.get() if self.payment_method else ''
            tk.Label(info_frame, text=f"Payment: {payment_method}",
                    bg='white', font=('Courier', 10)).pack(anchor='w')

        # Separator
        tk.Label(receipt, text="="*50, bg='white').pack()

        # Items
        text = tk.Text(receipt, height=12, width=45, bg='white', font=('Courier', 9))
        text.pack(padx=20, pady=10)

        # Insert items
        for item in self.bill_items:
            text.insert('end', f"{item['name'][:25]:25}\n")
            text.insert('end', f"  {item['qty']:3} x {item['price']:7.2f} = {item['total']:8.2f}\n")

        # Totals
        subtotal = sum(item['total'] for item in self.bill_items)
        tax_rate = float(self.tax_var.get())
        tax_amount = subtotal * (tax_rate/100)
        total = subtotal + tax_amount

        text.insert('end', "\n" + "-"*40 + "\n")
        text.insert('end', f"Subtotal: {subtotal:27.2f}\n")
        text.insert('end', f"VAT ({tax_rate}%): {tax_amount:24.2f}\n")
        text.insert('end', "-"*40 + "\n")
        text.insert('end', f"TOTAL: {total:30.2f} Birr\n")
        text.insert('end', "="*40 + "\n")
        text.insert('end', "Thank you for your purchase!\n")

        text.config(state='disabled')

        # Buttons
        btn_frame = tk.Frame(receipt, bg='white')
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="💾 Save PDF", bg='#3498db', fg='white',
                font=('Segoe UI', 10), command=lambda: self.save_receipt_pdf(text.get(1.0, 'end-1c'))).pack(side='left', padx=5)

        tk.Button(btn_frame, text="📄 Save Text", bg='#2ecc71', fg='white',
                font=('Segoe UI', 10), command=lambda: self.save_receipt_text(text.get(1.0, 'end-1c'))).pack(side='left', padx=5)

        tk.Button(btn_frame, text="🖨️ Print", bg='#e67e22', fg='white',
                font=('Segoe UI', 10), command=lambda: self.print_receipt(text.get(1.0, 'end-1c'))).pack(side='left', padx=5)

        tk.Button(btn_frame, text="❌ Close", bg='#e74c3c', fg='white',
                font=('Segoe UI', 10), command=receipt.destroy).pack(side='left', padx=5)
    
    def save_receipt_pdf(self, receipt_text):
        """Save receipt as PDF"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            receipts_dir = self.get_receipts_dir()
            filename = f"receipt_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = os.path.join(receipts_dir, filename)
            
            doc = SimpleDocTemplate(filepath, pagesize=A4,
                                   rightMargin=72, leftMargin=72,
                                   topMargin=72, bottomMargin=18)
            
            styles = getSampleStyleSheet()
            story = []
            
            # Header
            header_style = ParagraphStyle('CustomHeader', parent=styles['Heading1'],
                                         fontSize=16, textColor=colors.HexColor('#2c3e50'),
                                         spaceAfter=30, alignment=1)
            story.append(Paragraph("MERawi PHARMACY PRO", header_style))
            story.append(Paragraph("Tax Invoice", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            # Content
            content_style = ParagraphStyle('CustomBody', parent=styles['Normal'],
                                          fontSize=10, spaceAfter=6)
            
            for line in receipt_text.split('\n'):
                if line.strip():
                    story.append(Paragraph(line.replace(' ', '&nbsp;'), content_style))
            
            doc.build(story)
            
            messagebox.showinfo("✅ PDF Saved", f"Receipt saved as PDF:\n{filepath}")
            
            if messagebox.askyesno("Open Folder", "Open receipts folder?"):
                if platform.system() == "Windows":
                    os.startfile(receipts_dir)
                else:
                    os.system(f'open "{receipts_dir}"')
                    
        except ImportError:
            messagebox.showwarning("PDF Library Missing", 
                                 "PDF library not installed. Saving as text instead.\n\n"
                                 "To enable PDF: pip install reportlab")
            self.save_receipt_text(receipt_text)
        except Exception as e:
            messagebox.showerror("Error", f"Could not save PDF: {str(e)}")
    
    def save_receipt_text(self, receipt_text):
        """Save receipt as text"""
        receipts_dir = self.get_receipts_dir()
        filename = f"receipt_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        filepath = os.path.join(receipts_dir, filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(receipt_text)
        
        messagebox.showinfo("✅ Receipt Saved", f"Receipt saved:\n{filepath}")
        
        if messagebox.askyesno("Open Folder", "Open receipts folder?"):
            if platform.system() == "Windows":
                os.startfile(receipts_dir)
            else:
                os.system(f'open "{receipts_dir}"')
    
    def print_receipt(self, receipt_text):
        """Print receipt"""
        try:
            import tempfile
            import subprocess
            
            with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8') as f:
                f.write(receipt_text)
                temp_file = f.name
            
            if platform.system() == "Windows":
                os.startfile(temp_file, "print")
            elif platform.system() == "Darwin":
                subprocess.run(['lp', temp_file])
            else:
                subprocess.run(['lp', temp_file])
            
            messagebox.showinfo("🖨️ Printing", "Receipt sent to printer.")
            
        except Exception as e:
            messagebox.showerror("Print Error", f"Could not print: {str(e)}")
    
    # ========== REPORTS TAB ==========
    
    def show_invoice_list(self):
        """Window to manage all invoices"""
        conn = sqlite3.connect(self.get_database_path())
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, invoice_number, customer_name, total, amount_paid, due_date, status 
            FROM invoices ORDER BY date_issued DESC
        ''')
        invoices = cursor.fetchall()
        conn.close()
        
        win = tk.Toplevel(self.root)
        win.title("Invoice Management")
        win.geometry("900x500")
        
        tree = ttk.Treeview(win, columns=('ID', 'Number', 'Customer', 'Total', 'Paid', 'Due', 'Status'), show='headings')
        for col in ('ID', 'Number', 'Customer', 'Total', 'Paid', 'Due', 'Status'):
            tree.heading(col, text=col)
            tree.column(col, width=100)
        tree.column('Customer', width=200)
        tree.pack(fill='both', expand=True)
        
        for inv in invoices:
            tree.insert('', 'end', values=inv)
        
        def view_invoice():
            sel = tree.selection()
            if not sel: return
            inv_id = tree.item(sel[0])['values'][0]
            # Fetch full invoice details and show dialog with option to record payment or reprint
            self.view_invoice_details(inv_id)
        
        tk.Button(win, text="View / Pay", command=view_invoice, bg='#3498db', fg='white').pack(pady=10)

    def view_invoice_details(self, invoice_id):
        """Show invoice details, allow payment recording and reprint"""
        conn = sqlite3.connect(self.get_database_path())
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM invoices WHERE id = ?', (invoice_id,))
        inv = cursor.fetchone()
        cursor.execute('SELECT medicine_name, quantity, price_per_unit, total FROM invoice_items WHERE invoice_id = ?', (invoice_id,))
        items = cursor.fetchall()
        conn.close()
        
        # Create dialog
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Invoice {inv[1]}")
        dialog.geometry("600x500")
    
   
    
    def create_reports_tab(self):
        """Create reports tab"""
        self.reports_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.reports_frame, text="📊 Reports")
        
        # Buttons
        btn_frame = tk.Frame(self.reports_frame, bg='#ecf0f1', height=60)
        btn_frame.pack(fill='x')
        btn_frame.pack_propagate(False)
        
        reports = [
            ("⚠️ Low Stock", '#e67e22', self.low_stock_report),
            ("⏰ Expiring", '#f39c12', self.expiring_report),
            ("💰 Today's Sales", '#27ae60', self.today_sales_report),
            ("📋 All Medicines", '#2980b9', self.all_medicines_report)
        ]
        
        for text, color, cmd in reports:
            tk.Button(btn_frame, text=text, bg=color, fg='white',
                     font=('Segoe UI', 11), padx=20, command=cmd).pack(side='left', padx=10, pady=15)
        
        # Text area
        container = tk.Frame(self.reports_frame)
        container.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.report_text = tk.Text(container, height=25, font=('Courier', 10),
                                  wrap='word', bg='white', fg='#2c3e50')
        scrollbar = ttk.Scrollbar(container, orient='vertical', command=self.report_text.yview)
        self.report_text.configure(yscrollcommand=scrollbar.set)
        
        self.report_text.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
    
    def low_stock_report(self):
        """Low stock report"""
        self.notebook.select(self.reports_frame)
        self.report_text.config(state='normal')
        self.report_text.delete(1.0, tk.END)
        
        try:
            conn = sqlite3.connect(self.get_database_path())
            cursor = conn.cursor()
            cursor.execute('SELECT name, quantity, min_stock, location FROM medicines WHERE quantity <= min_stock ORDER BY quantity')
            items = cursor.fetchall()
            conn.close()
        except:
            items = []
        
        self.report_text.insert('end', "⚠️ LOW STOCK REPORT\n")
        self.report_text.insert('end', "="*60 + "\n\n")
        self.report_text.insert('end', f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n")
        
        if items:
            for item in items:
                self.report_text.insert('end', f"📦 {item[0]}\n")
                self.report_text.insert('end', f"   Stock: {item[1]} (Min: {item[2]})\n")
                self.report_text.insert('end', f"   Location: {item[3]}\n")
                self.report_text.insert('end', "-"*40 + "\n")
        else:
            self.report_text.insert('end', "✅ No low stock items\n")
        
        self.report_text.config(state='disabled')
    
    def expiring_report(self):
        """Expiring report"""
        self.notebook.select(self.reports_frame)
        self.report_text.config(state='normal')
        self.report_text.delete(1.0, tk.END)
        
        today = datetime.now().date()
        expiring = []
        
        try:
            conn = sqlite3.connect(self.get_database_path())
            cursor = conn.cursor()
            cursor.execute('SELECT name, expiry_date, quantity, location FROM medicines WHERE expiry_date IS NOT NULL')
            all_meds = cursor.fetchall()
            conn.close()
        except:
            all_meds = []
        
        for med in all_meds:
            if med[1] and len(med[1]) == 10:
                try:
                    expiry = datetime.strptime(med[1], '%Y-%m-%d').date()
                    days = (expiry - today).days
                    if 0 <= days <= 30:
                        expiring.append((med[0], med[1], days, med[2], med[3]))
                except:
                    pass
        
        expiring.sort(key=lambda x: x[2])
        
        self.report_text.insert('end', "⏰ EXPIRING MEDICINES (Next 30 Days)\n")
        self.report_text.insert('end', "="*60 + "\n\n")
        
        if expiring:
            for item in expiring:
                self.report_text.insert('end', f"💊 {item[0]}\n")
                self.report_text.insert('end', f"   Expires: {item[1]} (in {item[2]} days)\n")
                self.report_text.insert('end', f"   Quantity: {item[3]} at {item[4]}\n")
                self.report_text.insert('end', "-"*40 + "\n")
        else:
            self.report_text.insert('end', "✅ No medicines expiring soon\n")
        
        self.report_text.config(state='disabled')
    
    def today_sales_report(self):
        """Today's sales report"""
        self.notebook.select(self.reports_frame)
        self.report_text.config(state='normal')
        self.report_text.delete(1.0, tk.END)
        
        today = datetime.now().strftime('%Y-%m-%d')
        
        try:
            conn = sqlite3.connect(self.get_database_path())
            cursor = conn.cursor()
            cursor.execute('SELECT medicine_name, quantity, price_per_unit, total_price, sale_date FROM sales WHERE sale_date LIKE ? ORDER BY sale_date', (f'{today}%',))
            items = cursor.fetchall()
            
            cursor.execute('SELECT SUM(total_price) FROM sales WHERE sale_date LIKE ?', (f'{today}%',))
            total = cursor.fetchone()[0] or 0
            conn.close()
        except:
            items = []
            total = 0
        
        self.report_text.insert('end', f"💰 SALES REPORT FOR {today}\n")
        self.report_text.insert('end', "="*60 + "\n\n")
        
        if items:
            for item in items:
                self.report_text.insert('end', f"💊 {item[0]}\n")
                self.report_text.insert('end', f"   {item[1]} x {item[2]:.2f} = {item[3]:.2f}\n")
                self.report_text.insert('end', f"   Time: {item[4].split()[1] if ' ' in item[4] else item[4]}\n")
                self.report_text.insert('end', "-"*40 + "\n")
            
            self.report_text.insert('end', f"\n📊 TOTAL SALES: {total:.2f} Birr\n")
        else:
            self.report_text.insert('end', "No sales today\n")
        
        self.report_text.config(state='disabled')
    
    def all_medicines_report(self):
        """All medicines report"""
        self.notebook.select(self.reports_frame)
        self.report_text.config(state='normal')
        self.report_text.delete(1.0, tk.END)
        
        try:
            conn = sqlite3.connect(self.get_database_path())
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM medicines ORDER BY name')
            medicines = cursor.fetchall()
            conn.close()
        except:
            medicines = []
        
        self.report_text.insert('end', "📋 COMPLETE INVENTORY\n")
        self.report_text.insert('end', "="*80 + "\n\n")
        self.report_text.insert('end', f"Total: {len(medicines)} products\n\n")
        
        for med in medicines:
            self.report_text.insert('end', f"ID: {med[0]} | {med[1]}\n")
            self.report_text.insert('end', f"Batch: {med[4] or 'N/A'} | Exp: {med[5] or 'N/A'}\n")
            self.report_text.insert('end', f"Stock: {med[6]} | Price: {med[8]:.2f} | Loc: {med[10] or 'N/A'}\n")
            self.report_text.insert('end', "-"*50 + "\n")
        
        self.report_text.config(state='disabled')

    def refresh_dashboard(self):
        """Refresh dashboard data"""
        # Remove current dashboard tab
        for i in range(self.notebook.index('end')):
            if self.notebook.tab(i, 'text') == "📊 Dashboard":
                self.notebook.forget(i)
                break
        
        # Create new dashboard tab
        self.create_dashboard_tab()
        
        # Select dashboard tab
        for i in range(self.notebook.index('end')):
            if self.notebook.tab(i, 'text') == "📊 Dashboard":
                self.notebook.select(i)
                break 
   
    # ========== MEDICINE DIALOGS ==========
    
    def add_medicine_dialog(self):
        """Add medicine dialog"""
        dialog = tk.Toplevel(self.root)
        dialog.title("➕ Add Medicine")
        dialog.geometry("450x600")
        dialog.configure(bg='white')
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (450 // 2)
        y = (dialog.winfo_screenheight() // 2) - (600 // 2)
        dialog.geometry(f'+{x}+{y}')
        
        tk.Label(dialog, text="Add New Medicine", bg='#2c3e50', fg='white',
                font=('Segoe UI', 14, 'bold'), pady=10).pack(fill='x')
        
        main = tk.Frame(dialog, bg='white', padx=20, pady=20)
        main.pack(fill='both', expand=True)
        
        fields = [
            ("Medicine Name:", True),
            ("Category:", False),
            ("Manufacturer:", False),
            ("Batch Number:", True),
            ("Expiry Date (YYYY-MM-DD):", False),
            ("Quantity:", False),
            ("Purchase Price:", False),
            ("Selling Price:", False),
            ("Minimum Stock Alert:", False),
            ("Location:", False)
        ]
        
        entries = {}
        
        for label, required in fields:
            frame = tk.Frame(main, bg='white')
            frame.pack(fill='x', pady=5)
            
            text = label + (" *" if required else "")
            tk.Label(frame, text=text, bg='white', font=('Segoe UI', 10),
                    width=20, anchor='w').pack(side='left')
            
            entry = tk.Entry(frame, font=('Segoe UI', 10), width=25)
            entry.pack(side='left', padx=5)
            entries[label] = entry
        
        entries["Minimum Stock Alert:"].insert(0, "10")
        entries["Quantity:"].insert(0, "0")
        
        def save():
            name = entries["Medicine Name:"].get().strip()
            if not name:
                messagebox.showerror("Error", "Medicine name is required")
                return
            
            try:
                conn = sqlite3.connect(self.get_database_path())
                cursor = conn.cursor()
                
                cursor.execute('''
                    INSERT INTO medicines (
                        name, category, manufacturer, batch_number, expiry_date,
                        quantity, purchase_price, selling_price, min_stock, location, date_added
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    name,
                    entries["Category:"].get().strip(),
                    entries["Manufacturer:"].get().strip(),
                    entries["Batch Number:"].get().strip(),
                    entries["Expiry Date (YYYY-MM-DD):"].get().strip(),
                    int(entries["Quantity:"].get() or 0),
                    float(entries["Purchase Price:"].get() or 0),
                    float(entries["Selling Price:"].get() or 0),
                    int(entries["Minimum Stock Alert:"].get() or 10),
                    entries["Location:"].get().strip(),
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ))
                medicine_id = cursor.lastrowid
                quantity = int(entries["Quantity:"].get() or 0)
                if quantity > 0:
                    cursor.execute('''
                        INSERT INTO stock_movements (medicine_id, date, type, quantity, balance_before, balance_after, notes)
                        VALUES (?, ?, 'purchase', ?, 0, ?, ?)
                    ''', (medicine_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), quantity, quantity, "Initial stock"))
                conn.commit()
                conn.close()
                
                dialog.destroy()
                self.load_medicines()
                messagebox.showinfo("Success", "Medicine added successfully!")
                
            except Exception as e:
                messagebox.showerror("Error", f"Could not save: {str(e)}")
        
        # Buttons
        btn_frame = tk.Frame(dialog, bg='white')
        btn_frame.pack(pady=20)
        
        tk.Button(btn_frame, text="Save", bg='#27ae60', fg='white',
                 font=('Segoe UI', 12), padx=30, command=save).pack(side='left', padx=10)
        tk.Button(btn_frame, text="Cancel", bg='#e74c3c', fg='white',
                 font=('Segoe UI', 12), padx=30, command=dialog.destroy).pack(side='left', padx=10)
    
    def edit_medicine_dialog(self):
        """Edit medicine dialog"""
        selected = self.medicines_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a medicine")
            return
        
        item = self.medicines_tree.item(selected[0])
        med_id = item['values'][0]
        
        try:
            conn = sqlite3.connect(self.get_database_path())
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM medicines WHERE id = ?', (med_id,))
            med = cursor.fetchone()
            conn.close()
        except:
            med = None
        
        if not med:
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title("✏️ Edit Medicine")
        dialog.geometry("450x600")
        dialog.configure(bg='white')
        dialog.transient(self.root)
        dialog.grab_set()
        
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (450 // 2)
        y = (dialog.winfo_screenheight() // 2) - (600 // 2)
        dialog.geometry(f'+{x}+{y}')
        
        tk.Label(dialog, text="Edit Medicine", bg='#2980b9', fg='white',
                font=('Segoe UI', 14, 'bold'), pady=10).pack(fill='x')
        
        main = tk.Frame(dialog, bg='white', padx=20, pady=20)
        main.pack(fill='both', expand=True)
        
        fields = [
            ("Medicine Name:", med[1]),
            ("Category:", med[2] or ""),
            ("Manufacturer:", med[3] or ""),
            ("Batch Number:", med[4] or ""),
            ("Expiry Date:", med[5] or ""),
            ("Quantity:", med[6]),
            ("Purchase Price:", med[7]),
            ("Selling Price:", med[8]),
            ("Minimum Stock Alert:", med[9]),
            ("Location:", med[10] or "")
        ]
        
        entries = {}
        
        for label, value in fields:
            frame = tk.Frame(main, bg='white')
            frame.pack(fill='x', pady=5)
            
            tk.Label(frame, text=label, bg='white', font=('Segoe UI', 10),
                    width=20, anchor='w').pack(side='left')
            
            entry = tk.Entry(frame, font=('Segoe UI', 10), width=25)
            entry.pack(side='left', padx=5)
            entry.insert(0, str(value))
            entries[label] = entry
        
        def update():
            try:
                conn = sqlite3.connect(self.get_database_path())
                cursor = conn.cursor()
                cursor.execute('''
                    UPDATE medicines SET
                        name = ?, category = ?, manufacturer = ?, batch_number = ?,
                        expiry_date = ?, quantity = ?, purchase_price = ?,
                        selling_price = ?, min_stock = ?, location = ?
                    WHERE id = ?
                ''', (
                    entries["Medicine Name:"].get(),
                    entries["Category:"].get(),
                    entries["Manufacturer:"].get(),
                    entries["Batch Number:"].get(),
                    entries["Expiry Date:"].get(),
                    int(entries["Quantity:"].get() or 0),
                    float(entries["Purchase Price:"].get() or 0),
                    float(entries["Selling Price:"].get() or 0),
                    int(entries["Minimum Stock Alert:"].get() or 10),
                    entries["Location:"].get(),
                    med_id
                ))
                old_qty = med[6]  
                new_qty = int(entries["Quantity:"].get() or 0)
                diff = new_qty - old_qty
                # After updating the medicine:
                if diff != 0:
                    cursor.execute('''
                        INSERT INTO stock_movements (medicine_id, date, type, quantity, balance_before, balance_after, notes)
                        VALUES (?, ?, 'adjustment', ?, ?, ?, ?)
                    ''', (med_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), diff, old_qty, new_qty, "Manual adjustment"))
                conn.commit()
                conn.close()
                dialog.destroy()
                self.load_medicines()
                messagebox.showinfo("Success", "Medicine updated!")
                
            except Exception as e:
                messagebox.showerror("Error", f"Update failed: {str(e)}")
        
        btn_frame = tk.Frame(dialog, bg='white')
        btn_frame.pack(pady=20)
        
        tk.Button(btn_frame, text="Update", bg='#2980b9', fg='white',
                 font=('Segoe UI', 12), padx=30, command=update).pack(side='left', padx=10)
        tk.Button(btn_frame, text="Cancel", bg='#e74c3c', fg='white',
                 font=('Segoe UI', 12), padx=30, command=dialog.destroy).pack(side='left', padx=10)
    
    def delete_medicine(self):
        """Delete medicine and record deletion movement"""
        selected = self.medicines_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a medicine")
            return
        
        if not messagebox.askyesno("Confirm", "Delete this medicine? This will remove all stock records."):
            return
        
        item = self.medicines_tree.item(selected[0])
        med_id = item['values'][0]
        
        try:
            conn = sqlite3.connect(self.get_database_path())
            cursor = conn.cursor()
            
            # Get current quantity before deletion
            cursor.execute('SELECT quantity FROM medicines WHERE id = ?', (med_id,))
            old_qty = cursor.fetchone()[0]
            
            # Insert stock movement to zero out
            if old_qty > 0:
                cursor.execute('''
                    INSERT INTO stock_movements (medicine_id, date, type, quantity, balance_before, balance_after, notes)
                    VALUES (?, ?, 'deletion', ?, ?, 0, ?)
                ''', (med_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), -old_qty, old_qty, "Medicine deleted"))
            
            # Delete the medicine
            cursor.execute('DELETE FROM medicines WHERE id = ?', (med_id,))
            
            conn.commit()
            conn.close()
            
            self.load_medicines()
            messagebox.showinfo("Success", "Medicine deleted")
            
        except Exception as e:
            messagebox.showerror("Error", f"Delete failed: {str(e)}")
        
    # ========== BACKUP & IMPORT ==========
    
    def backup_to_excel(self):
        """Backup to Excel"""
        backups_dir = self.get_backups_dir()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        try:
            conn = sqlite3.connect(self.get_database_path())
            cursor = conn.cursor()
            
            # Medicines
            cursor.execute('SELECT * FROM medicines')
            medicines = cursor.fetchall()
            
            med_file = os.path.join(backups_dir, f'medicines_{timestamp}.csv')
            with open(med_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['ID', 'Name', 'Category', 'Manufacturer', 'Batch', 'Expiry',
                               'Quantity', 'Purchase Price', 'Selling Price', 'Min Stock', 'Location', 'Date Added'])
                writer.writerows(medicines)
            
            # Sales
            cursor.execute('SELECT * FROM sales')
            sales = cursor.fetchall()
            
            sales_file = os.path.join(backups_dir, f'sales_{timestamp}.csv')
            with open(sales_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['ID', 'Medicine ID', 'Medicine Name', 'Quantity', 'Price/Unit',
                               'Total', 'Sale Date', 'Customer'])
                writer.writerows(sales)
            
            conn.close()
            
            messagebox.showinfo("✅ Backup Complete",
                              f"Backup saved to:\n{backups_dir}\n\n"
                              f"Files:\n{os.path.basename(med_file)}\n{os.path.basename(sales_file)}")
            
            if messagebox.askyesno("Open Folder", "Open backups folder?"):
                if platform.system() == "Windows":
                    os.startfile(backups_dir)
                else:
                    os.system(f'open "{backups_dir}"')
                    
        except Exception as e:
            messagebox.showerror("Error", f"Backup failed: {str(e)}")
    
    def import_from_excel(self):
        """Import from Excel with header validation and bin card recording (using name, category, batch as key)"""
        filename = filedialog.askopenfilename(
            title="Select File",
            filetypes=[("Excel files", "*.csv *.xlsx"), ("All files", "*.*")]
        )
        
        if not filename:
            return
        
        try:
            # Save copy
            exports_dir = self.get_exports_dir()
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            import_copy = os.path.join(exports_dir, f"imported_{timestamp}_{os.path.basename(filename)}")
            shutil.copy2(filename, import_copy)
            
            conn = sqlite3.connect(self.get_database_path())
            cursor = conn.cursor()
            
            import_count = 0
            update_count = 0
            skip_count = 0
            error_rows = []
            
            # Expected columns (in order, but we map by name)
            expected_columns = [
                "name", "category", "manufacturer", "batch_number", "expiry_date",
                "quantity", "purchase_price", "selling_price", "min_stock", "location"
            ]
            # Required columns for validation
            required_columns = ["name", "category", "batch_number", "quantity", "selling_price"]
            
            def map_header(header_row):
                """Return dict mapping expected column name to index using flexible aliases."""
                # Aliases for each expected column
                aliases = {
                    "name": ["name", "medicine", "product"],
                    "category": ["category", "cat"],
                    "manufacturer": ["manufacturer", "company", "brand"],
                    "batch_number": ["batch_number", "batch", "batchno", "lot"],
                    "expiry_date": ["expiry_date", "expiry", "exp", "expiration"],
                    "quantity": ["quantity", "qty", "stock"],
                    "purchase_price": ["purchase_price", "purchase", "cost", "buy price"],
                    "selling_price": ["selling_price", "selling", "price", "sell price"],
                    "min_stock": ["min_stock", "min", "minimum", "alert"],
                    "location": ["location", "shelf", "rack"]
                }
                header_lower = [str(cell).strip().lower() for cell in header_row]
                col_map = {}
                for expected, alias_list in aliases.items():
                    for i, col in enumerate(header_lower):
                        if any(alias in col for alias in alias_list):
                            col_map[expected] = i
                            break
                # Check required columns
                missing = [col for col in required_columns if col not in col_map]
                if missing:
                    raise ValueError(f"Missing required columns: {', '.join(missing)}")
                return col_map
            
            # ---------- CSV ----------
            if filename.lower().endswith('.csv'):
                with open(filename, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    header = next(reader, None)
                    if not header:
                        raise ValueError("File is empty")
                    col_map = map_header(header)
                    
                    for row_num, row in enumerate(reader, start=2):
                        # Ensure row has enough columns
                        if len(row) < max(col_map.values()) + 1:
                            skip_count += 1
                            error_rows.append(f"Row {row_num}: insufficient columns")
                            continue
                        
                        try:
                            # Required fields
                            name = row[col_map["name"]].strip()
                            if not name:
                                skip_count += 1
                                error_rows.append(f"Row {row_num}: missing name")
                                continue
                            
                            category = row[col_map["category"]].strip()
                            if not category:
                                skip_count += 1
                                error_rows.append(f"Row {row_num}: missing category")
                                continue
                            
                            batch = row[col_map["batch_number"]].strip()
                            if not batch:
                                skip_count += 1
                                error_rows.append(f"Row {row_num}: missing batch number")
                                continue
                            
                            # Numeric required
                            qty_str = row[col_map["quantity"]].strip()
                            try:
                                qty = int(float(qty_str)) if qty_str else 0
                            except:
                                qty = 0
                            
                            selling_str = row[col_map["selling_price"]].strip()
                            try:
                                selling = float(selling_str) if selling_str else 0
                            except:
                                selling = 0
                            
                            # Optional fields
                            manufacturer = row[col_map["manufacturer"]].strip() if "manufacturer" in col_map else ""
                            expiry = row[col_map["expiry_date"]].strip() if "expiry_date" in col_map else ""
                            purchase_str = row[col_map["purchase_price"]].strip() if "purchase_price" in col_map else "0"
                            try:
                                purchase = float(purchase_str) if purchase_str else 0
                            except:
                                purchase = 0
                            min_stock_str = row[col_map["min_stock"]].strip() if "min_stock" in col_map else "10"
                            try:
                                min_stock = int(float(min_stock_str)) if min_stock_str else 10
                            except:
                                min_stock = 10
                            location = row[col_map["location"]].strip() if "location" in col_map else ""
                            
                            # Check if medicine exists by name, category, batch
                            cursor.execute('''
                                SELECT id, quantity FROM medicines 
                                WHERE name = ? AND category = ? AND batch_number = ?
                            ''', (name, category, batch))
                            existing = cursor.fetchone()
                            
                            if existing:
                                med_id, old_qty = existing
                                new_qty = old_qty + qty
                                # Update medicine (update all fields except id and date_added)
                                cursor.execute('''
                                    UPDATE medicines SET
                                        manufacturer = ?,
                                        expiry_date = ?,
                                        quantity = ?,
                                        purchase_price = ?,
                                        selling_price = ?,
                                        min_stock = ?,
                                        location = ?
                                    WHERE id = ?
                                ''', (manufacturer, expiry, new_qty, purchase, selling, min_stock, location, med_id))
                                
                                # Record bin movement if quantity increased
                                if qty > 0:
                                    cursor.execute('''
                                        INSERT INTO stock_movements (medicine_id, date, type, quantity, balance_before, balance_after, notes)
                                        VALUES (?, ?, 'adjustment', ?, ?, ?, ?)
                                    ''', (med_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), qty, old_qty, new_qty, f"Import adjustment (+{qty})"))
                                update_count += 1
                            else:
                                # Insert new medicine
                                cursor.execute('''
                                    INSERT INTO medicines (
                                        name, category, manufacturer, batch_number, expiry_date,
                                        quantity, purchase_price, selling_price, min_stock, location, date_added
                                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                ''', (name, category, manufacturer, batch, expiry, qty, purchase,
                                    selling, min_stock, location, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                                med_id = cursor.lastrowid
                                
                                # Record initial stock movement
                                if qty > 0:
                                    cursor.execute('''
                                        INSERT INTO stock_movements (medicine_id, date, type, quantity, balance_before, balance_after, notes)
                                        VALUES (?, ?, 'purchase', ?, 0, ?, ?)
                                    ''', (med_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), qty, qty, "Initial stock from import"))
                                import_count += 1
                                
                        except Exception as e:
                            skip_count += 1
                            error_rows.append(f"Row {row_num}: {str(e)}")
                            continue
            
            # ---------- Excel (XLSX) ----------
            elif filename.lower().endswith('.xlsx'):
                from openpyxl import load_workbook
                wb = load_workbook(filename, data_only=True)
                sheet = wb.active
                
                header_row = [cell.value for cell in sheet[1]]
                if not header_row:
                    raise ValueError("File is empty")
                col_map = map_header(header_row)
                
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not row:
                        continue
                    
                    try:
                        # Required fields
                        name = str(row[col_map["name"]]).strip() if row[col_map["name"]] else ""
                        if not name:
                            skip_count += 1
                            error_rows.append(f"Row {row_num}: missing name")
                            continue
                        
                        category = str(row[col_map["category"]]).strip() if row[col_map["category"]] else ""
                        if not category:
                            skip_count += 1
                            error_rows.append(f"Row {row_num}: missing category")
                            continue
                        
                        batch = str(row[col_map["batch_number"]]).strip() if row[col_map["batch_number"]] else ""
                        if not batch:
                            skip_count += 1
                            error_rows.append(f"Row {row_num}: missing batch number")
                            continue
                        
                        # Numeric required
                        qty_val = row[col_map["quantity"]]
                        try:
                            qty = int(float(qty_val)) if qty_val is not None else 0
                        except:
                            qty = 0
                        
                        selling_val = row[col_map["selling_price"]]
                        try:
                            selling = float(selling_val) if selling_val is not None else 0
                        except:
                            selling = 0
                        
                        # Optional
                        manufacturer = str(row[col_map["manufacturer"]]).strip() if "manufacturer" in col_map and row[col_map["manufacturer"]] else ""
                        expiry = row[col_map["expiry_date"]] if "expiry_date" in col_map else ""
                        if expiry and isinstance(expiry, datetime):
                            expiry = expiry.strftime('%Y-%m-%d')
                        elif expiry:
                            expiry = str(expiry).strip()
                        else:
                            expiry = ""
                        
                        purchase_val = row[col_map["purchase_price"]] if "purchase_price" in col_map else 0
                        try:
                            purchase = float(purchase_val) if purchase_val is not None else 0
                        except:
                            purchase = 0
                        
                        min_stock_val = row[col_map["min_stock"]] if "min_stock" in col_map else 10
                        try:
                            min_stock = int(float(min_stock_val)) if min_stock_val is not None else 10
                        except:
                            min_stock = 10
                        
                        location = str(row[col_map["location"]]).strip() if "location" in col_map and row[col_map["location"]] else ""
                        
                        # Check if exists
                        cursor.execute('''
                            SELECT id, quantity FROM medicines 
                            WHERE name = ? AND category = ? AND batch_number = ?
                        ''', (name, category, batch))
                        existing = cursor.fetchone()
                        
                        if existing:
                            med_id, old_qty = existing
                            new_qty = old_qty + qty
                            cursor.execute('''
                                UPDATE medicines SET
                                    manufacturer = ?,
                                    expiry_date = ?,
                                    quantity = ?,
                                    purchase_price = ?,
                                    selling_price = ?,
                                    min_stock = ?,
                                    location = ?
                                WHERE id = ?
                            ''', (manufacturer, expiry, new_qty, purchase, selling, min_stock, location, med_id))
                            
                            if qty > 0:
                                cursor.execute('''
                                    INSERT INTO stock_movements (medicine_id, date, type, quantity, balance_before, balance_after, notes)
                                    VALUES (?, ?, 'adjustment', ?, ?, ?, ?)
                                ''', (med_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), qty, old_qty, new_qty, f"Import adjustment (+{qty})"))
                            update_count += 1
                        else:
                            cursor.execute('''
                                INSERT INTO medicines (
                                    name, category, manufacturer, batch_number, expiry_date,
                                    quantity, purchase_price, selling_price, min_stock, location, date_added
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ''', (name, category, manufacturer, batch, expiry, qty, purchase,
                                selling, min_stock, location, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                            med_id = cursor.lastrowid
                            
                            if qty > 0:
                                cursor.execute('''
                                    INSERT INTO stock_movements (medicine_id, date, type, quantity, balance_before, balance_after, notes)
                                    VALUES (?, ?, 'purchase', ?, 0, ?, ?)
                                ''', (med_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), qty, qty, "Initial stock from import"))
                            import_count += 1
                            
                    except Exception as e:
                        skip_count += 1
                        error_rows.append(f"Row {row_num}: {str(e)}")
                        continue
            
            else:
                messagebox.showerror("Error", "Unsupported file type")
                return
            
            # Commit all changes
            conn.commit()
            cursor.execute('SELECT COUNT(*) FROM medicines')
            total = cursor.fetchone()[0]
            conn.close()
            
            self.load_medicines()
            
            # Build result message
            message = f"✅ IMPORT COMPLETE\n\n"
            message += f"File: {os.path.basename(filename)}\n"
            message += f"New medicines added: {import_count}\n"
            message += f"Existing medicines updated: {update_count}\n"
            message += f"Rows skipped: {skip_count}\n"
            message += f"Total medicines now: {total}\n\n"
            if error_rows:
                message += f"Errors (first 5):\n" + "\n".join(error_rows[:5]) + "\n\n"
            message += f"Copy saved to:\n{import_copy}"
            
            messagebox.showinfo("Import Results", message)
            
        except Exception as e:
            messagebox.showerror("Import Error", f"Failed to import:\n{str(e)}")
    # ========== ADVANCED DATABASE ANALYSIS ==========
    
    def test_database_contents(self):
        """PREMIUM BUSINESS INTELLIGENCE DASHBOARD"""
        conn = sqlite3.connect(self.get_database_path())
        cursor = conn.cursor()
        
        # Basic metrics
        cursor.execute('SELECT COUNT(*) FROM medicines')
        total_products = cursor.fetchone()[0]
        
        cursor.execute('SELECT SUM(quantity) FROM medicines')
        total_units = cursor.fetchone()[0] or 0
        
        cursor.execute('SELECT SUM(quantity * selling_price) FROM medicines')
        total_value = cursor.fetchone()[0] or 0
        
        cursor.execute('SELECT SUM(quantity * purchase_price) FROM medicines')
        total_cost = cursor.fetchone()[0] or 0
        
        potential_profit = total_value - total_cost
        profit_margin = (potential_profit / total_value * 100) if total_value > 0 else 0
        
        # Sales
        cursor.execute('SELECT SUM(total_price), COUNT(*) FROM sales')
        sales = cursor.fetchone()
        total_revenue = sales[0] or 0
        total_transactions = sales[1] or 0
        
        # Today
        today = datetime.now().strftime('%Y-%m-%d')
        cursor.execute('SELECT SUM(total_price), COUNT(*) FROM sales WHERE sale_date LIKE ?', (f'{today}%',))
        today_data = cursor.fetchone()
        today_revenue = today_data[0] or 0
        today_transactions = today_data[1] or 0
        
        # Status analysis
        today_date = datetime.now().date()
        cursor.execute('SELECT expiry_date, quantity, min_stock, selling_price, purchase_price FROM medicines')
        all_meds = cursor.fetchall()
        
        status = {
            'healthy': {'count': 0, 'units': 0, 'value': 0},
            'low_stock': {'count': 0, 'units': 0, 'value': 0},
            'expiring': {'count': 0, 'units': 0, 'value': 0},
            'expired': {'count': 0, 'units': 0, 'value': 0, 'loss': 0}
        }
        
        for expiry_str, qty, min_stock, sell_price, buy_price in all_meds:
            item_value = qty * sell_price
            item_cost = qty * buy_price
            is_expired = False
            is_expiring = False
            
            if expiry_str and len(expiry_str) == 10:
                try:
                    expiry = datetime.strptime(expiry_str, '%Y-%m-%d').date()
                    days = (expiry - today_date).days
                    
                    if days < 0:
                        status['expired']['count'] += 1
                        status['expired']['units'] += qty
                        status['expired']['value'] += item_value
                        status['expired']['loss'] += item_cost
                        is_expired = True
                    elif days <= 30:
                        status['expiring']['count'] += 1
                        status['expiring']['units'] += qty
                        status['expiring']['value'] += item_value
                        is_expiring = True
                except:
                    pass
            
            if not is_expired and not is_expiring:
                if qty <= min_stock:
                    status['low_stock']['count'] += 1
                    status['low_stock']['units'] += qty
                    status['low_stock']['value'] += item_value
                else:
                    status['healthy']['count'] += 1
                    status['healthy']['units'] += qty
                    status['healthy']['value'] += item_value
        
        # Categories
        cursor.execute('''
            SELECT category, COUNT(*) FROM medicines
            WHERE category IS NOT NULL AND category != ''
            GROUP BY category ORDER BY COUNT(*) DESC LIMIT 5
        ''')
        categories = cursor.fetchall()
        
        conn.close()
        
        # Turnover
        turnover = total_revenue / total_value if total_value > 0 else 0
        days_inventory = total_value / (total_revenue / 30) if total_revenue > 0 else 0
        
        # Create window
        dashboard = tk.Toplevel(self.root)
        dashboard.title("📊 BUSINESS INTELLIGENCE")
        dashboard.geometry("900x750")
        dashboard.configure(bg='#1a2634')
        dashboard.transient(self.root)
        dashboard.grab_set()
        
        # Center
        dashboard.update_idletasks()
        x = (dashboard.winfo_screenwidth() // 2) - (900 // 2)
        y = (dashboard.winfo_screenheight() // 2) - (750 // 2)
        dashboard.geometry(f'+{x}+{y}')
        
        # Scrollable
        canvas = tk.Canvas(dashboard, bg='#1a2634', highlightthickness=0)
        scrollbar = tk.Scrollbar(dashboard, orient='vertical', command=canvas.yview)
        scrollable = tk.Frame(canvas, bg='#1a2634')
        
        scrollable.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=scrollable, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        def wheel(e): canvas.yview_scroll(int(-1*(e.delta/120)), 'units')
        scrollable.bind('<Enter>', lambda e: canvas.bind_all('<MouseWheel>', wheel))
        scrollable.bind('<Leave>', lambda e: canvas.unbind_all('<MouseWheel>'))
        
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Main content
        main = tk.Frame(scrollable, bg='#1a2634', padx=25, pady=25)
        main.pack(fill='both', expand=True)
        
        # Header
        header = tk.Frame(main, bg='#1a2634')
        header.pack(fill='x', pady=(0, 20))
        
        tk.Label(header, text="⚕️ MERawi PHARMACY PRO", bg='#1a2634', fg='#f1c40f',
                font=('Segoe UI', 18, 'bold')).pack(anchor='w')
        tk.Label(header, text="Business Intelligence Dashboard", bg='#1a2634', fg='#95a5a6',
                font=('Segoe UI', 10)).pack(anchor='w')
        
        tk.Label(header, text=datetime.now().strftime('%B %d, %Y %I:%M %p'),
                bg='#1a2634', fg='#7f8c8d', font=('Segoe UI', 9)).pack(anchor='e')
        
        # KPI Row 1
        kpi1 = tk.Frame(main, bg='#1a2634')
        kpi1.pack(fill='x', pady=5)
        
        # Card 1
        card1 = tk.Frame(kpi1, bg='#2c3e50', padx=15, pady=15)
        card1.pack(side='left', fill='both', expand=True, padx=2)
        tk.Label(card1, text="TOTAL VALUE", bg='#2c3e50', fg='#bdc3c7',
                font=('Segoe UI', 8)).pack(anchor='w')
        tk.Label(card1, text=f"{total_value:,.0f}", bg='#2c3e50', fg='#2ecc71',
                font=('Segoe UI', 20, 'bold')).pack(anchor='w')
        tk.Label(card1, text="Birr", bg='#2c3e50', fg='#95a5a6',
                font=('Segoe UI', 8)).pack(anchor='w')
        
        # Card 2
        card2 = tk.Frame(kpi1, bg='#2c3e50', padx=15, pady=15)
        card2.pack(side='left', fill='both', expand=True, padx=2)
        tk.Label(card2, text="POTENTIAL PROFIT", bg='#2c3e50', fg='#bdc3c7',
                font=('Segoe UI', 8)).pack(anchor='w')
        tk.Label(card2, text=f"{potential_profit:,.0f}", bg='#2c3e50', fg='#f1c40f',
                font=('Segoe UI', 20, 'bold')).pack(anchor='w')
        tk.Label(card2, text=f"Margin: {profit_margin:.1f}%", bg='#2c3e50', fg='#95a5a6',
                font=('Segoe UI', 8)).pack(anchor='w')
        
        # Card 3
        card3 = tk.Frame(kpi1, bg='#2c3e50', padx=15, pady=15)
        card3.pack(side='left', fill='both', expand=True, padx=2)
        tk.Label(card3, text="ACTUAL REVENUE", bg='#2c3e50', fg='#bdc3c7',
                font=('Segoe UI', 8)).pack(anchor='w')
        tk.Label(card3, text=f"{total_revenue:,.0f}", bg='#2c3e50', fg='#3498db',
                font=('Segoe UI', 20, 'bold')).pack(anchor='w')
        tk.Label(card3, text=f"{total_transactions} sales", bg='#2c3e50', fg='#95a5a6',
                font=('Segoe UI', 8)).pack(anchor='w')
        
        # KPI Row 2
        kpi2 = tk.Frame(main, bg='#1a2634')
        kpi2.pack(fill='x', pady=5)
        
        # Card 4
        card4 = tk.Frame(kpi2, bg='#2c3e50', padx=15, pady=15)
        card4.pack(side='left', fill='both', expand=True, padx=2)
        tk.Label(card4, text="TODAY'S SALES", bg='#2c3e50', fg='#bdc3c7',
                font=('Segoe UI', 8)).pack(anchor='w')
        tk.Label(card4, text=f"{today_revenue:,.0f}", bg='#2c3e50', fg='#2ecc71',
                font=('Segoe UI', 18, 'bold')).pack(side='left', padx=5)
        tk.Label(card4, text=f"transactions({today_transactions})", bg='#2c3e50', fg='#95a5a6',
                font=('Segoe UI', 10)).pack(side='left')
        
        # Card 5
        card5 = tk.Frame(kpi2, bg='#2c3e50', padx=15, pady=15)
        card5.pack(side='left', fill='both', expand=True, padx=2)
        tk.Label(card5, text="TURNOVER RATE", bg='#2c3e50', fg='#bdc3c7',
                font=('Segoe UI', 8)).pack(anchor='w')
        color = '#2ecc71' if turnover >= 4 else '#f39c12' if turnover >= 2 else '#e74c3c'
        tk.Label(card5, text=f"{turnover:.2f}x", bg='#2c3e50', fg=color,
                font=('Segoe UI', 18, 'bold')).pack(anchor='w')
        
        # Card 6
        card6 = tk.Frame(kpi2, bg='#2c3e50', padx=15, pady=15)
        card6.pack(side='left', fill='both', expand=True, padx=2)
        tk.Label(card6, text="INVENTORY DAYS", bg='#2c3e50', fg='#bdc3c7',
                font=('Segoe UI', 8)).pack(anchor='w')
        days_color = '#e74c3c' if days_inventory > 60 else '#f39c12' if days_inventory > 30 else '#2ecc71'
        tk.Label(card6, text=f"{days_inventory:.0f}", bg='#2c3e50', fg=days_color,
                font=('Segoe UI', 18, 'bold')).pack(anchor='w')
        
        # Status section
        status_frame = tk.Frame(main, bg='#1a2634')
        status_frame.pack(fill='x', pady=15)
        
        tk.Label(status_frame, text="📊 INVENTORY HEALTH", bg='#1a2634', fg='#ecf0f1',
                font=('Segoe UI', 12, 'bold')).pack(anchor='w', pady=(0, 10))
        
        health_items = [
            ("✅ HEALTHY", status['healthy']['count'], status['healthy']['units'], status['healthy']['value'], '#2ecc71'),
            ("⚠️ LOW STOCK", status['low_stock']['count'], status['low_stock']['units'], status['low_stock']['value'], '#f39c12'),
            ("⏰ EXPIRING", status['expiring']['count'], status['expiring']['units'], status['expiring']['value'], '#e67e22'),
            ("❌ EXPIRED", status['expired']['count'], status['expired']['units'], status['expired']['value'], '#e74c3c')
        ]
        
        for label, count, units, value, color in health_items:
            item = tk.Frame(status_frame, bg='#2c3e50', padx=15, pady=8)
            item.pack(fill='x', pady=2)
            
            tk.Label(item, text=label, bg='#2c3e50', fg=color,
                    font=('Segoe UI', 10, 'bold'), width=14, anchor='w').pack(side='left')
            tk.Label(item, text=f"{count} products", bg='#2c3e50', fg='#ecf0f1',
                    font=('Segoe UI', 10), width=10).pack(side='left')
            tk.Label(item, text=f"{units:,} units", bg='#2c3e50', fg='#bdc3c7',
                    font=('Segoe UI', 10), width=12).pack(side='left')
            tk.Label(item, text=f"{value:,.0f} Birr", bg='#2c3e50', fg='#f1c40f',
                    font=('Segoe UI', 10, 'bold')).pack(side='right')
        
        # Loss analysis
        loss_frame = tk.Frame(main, bg='#1a2634')
        loss_frame.pack(fill='x', pady=15)
        
        tk.Label(loss_frame, text="💰 RISK ASSESSMENT", bg='#1a2634', fg='#ecf0f1',
                font=('Segoe UI', 12, 'bold')).pack(anchor='w', pady=(0, 10))
        
        loss_container = tk.Frame(loss_frame, bg='#2c3e50', padx=20, pady=15)
        loss_container.pack(fill='x')
        
        rows = [
            ("Cash in Inventory:", f"{total_cost:,.0f} Birr", '#3498db'),
            ("Expired Loss:", f"{status['expired']['loss']:,.0f} Birr", '#e74c3c'),
            ("At Risk (Expiring):", f"{status['expiring']['value'] * 0.3:,.0f} Birr", '#f39c12'),
        ]
        
        for label, value, color in rows:
            row = tk.Frame(loss_container, bg='#2c3e50')
            row.pack(fill='x', pady=3)
            tk.Label(row, text=label, bg='#2c3e50', fg='#bdc3c7',
                    font=('Segoe UI', 10)).pack(side='left')
            tk.Label(row, text=value, bg='#2c3e50', fg=color,
                    font=('Segoe UI', 11, 'bold')).pack(side='right')
        
        total_loss = status['expired']['loss'] + (status['expiring']['value'] * 0.3)
        tk.Frame(loss_container, bg='#34495e', height=1).pack(fill='x', pady=10)
        
        total_row = tk.Frame(loss_container, bg='#2c3e50')
        total_row.pack(fill='x')
        tk.Label(total_row, text="TOTAL POTENTIAL LOSS:", bg='#2c3e50', fg='#ecf0f1',
                font=('Segoe UI', 11, 'bold')).pack(side='left')
        tk.Label(total_row, text=f"{total_loss:,.0f} Birr", bg='#2c3e50', fg='#e74c3c',
                font=('Segoe UI', 13, 'bold')).pack(side='right')
        
        # Recommendations
        recommendations = []
        if status['low_stock']['count'] > 0:
            recommendations.append(f"🔴 Restock {status['low_stock']['count']} low items")
        if status['expiring']['count'] > 0:
            recommendations.append(f"🟡 Run sale on {status['expiring']['count']} expiring items")
        if status['expired']['count'] > 0:
            recommendations.append(f"❌ Remove {status['expired']['count']} expired items")
        if turnover < 2:
            recommendations.append("📉 Slow turnover - consider promotions")
        elif turnover > 6:
            recommendations.append("📈 High demand - increase stock")
        
        if recommendations:
            rec_frame = tk.Frame(main, bg='#1a2634')
            rec_frame.pack(fill='x', pady=15)
            
            tk.Label(rec_frame, text="🎯 RECOMMENDATIONS", bg='#1a2634', fg='#f1c40f',
                    font=('Segoe UI', 12, 'bold')).pack(anchor='w', pady=(0, 10))
            
            for rec in recommendations:
                r = tk.Frame(rec_frame, bg='#2c3e50', padx=15, pady=8)
                r.pack(fill='x', pady=2)
                tk.Label(r, text=rec, bg='#2c3e50', fg='#ecf0f1',
                        font=('Segoe UI', 10)).pack(anchor='w')
        
        # Executive summary
        summary_frame = tk.Frame(main, bg='#1a2634')
        summary_frame.pack(fill='x', pady=15)
        
        tk.Label(summary_frame, text="📋 EXECUTIVE SUMMARY", bg='#1a2634', fg='#ecf0f1',
                font=('Segoe UI', 12, 'bold')).pack(anchor='w', pady=(0, 10))
        
        summary_box = tk.Frame(summary_frame, bg='#2c3e50', padx=20, pady=15)
        summary_box.pack(fill='x')
        
        health_score = 100
        health_score -= status['expired']['count'] * 5
        health_score -= status['expiring']['count'] * 2
        health_score -= status['low_stock']['count'] * 3
        health_score = max(0, min(100, health_score))
        
        score_color = '#2ecc71' if health_score >= 70 else '#f39c12' if health_score >= 40 else '#e74c3c'
        
        score_row = tk.Frame(summary_box, bg='#2c3e50')
        score_row.pack(fill='x', pady=5)
        tk.Label(score_row, text="Health Score:", bg='#2c3e50', fg='#bdc3c7',
                font=('Segoe UI', 11)).pack(side='left')
        tk.Label(score_row, text=f"{health_score}/100", bg='#2c3e50', fg=score_color,
                font=('Segoe UI', 16, 'bold')).pack(side='right')
        
        summary = f"Your pharmacy has {total_products} products worth {total_value:,.0f} Birr. "
        summary += f"Profit margin is {profit_margin:.1f}%. "
        if status['expired']['loss'] > 0:
            summary += f"Lost {status['expired']['loss']:,.0f} Birr to expired items. "
        
        tk.Label(summary_box, text=summary, bg='#2c3e50', fg='#ecf0f1',
                font=('Segoe UI', 10), wraplength=700).pack(pady=10)
        
        # Footer
        footer = tk.Frame(main, bg='#1a2634')
        footer.pack(fill='x', pady=(20, 10))
        
        tk.Label(footer, text="MERawi PHARMACY PRO · Business Intelligence", 
                bg='#1a2634', fg='#5d6d7e', font=('Segoe UI', 8)).pack()
        tk.Label(footer, text="Merawi Yohannes · 0921-540-245 · merawiyohannes@gmail.com", 
                bg='#1a2634', fg='#5d6d7e', font=('Segoe UI', 8)).pack()
        
        # Close button
        tk.Button(main, text="CLOSE DASHBOARD", bg='#34495e', fg='white',
                 font=('Segoe UI', 10, 'bold'), padx=40, pady=8,
                 relief='flat', command=dashboard.destroy).pack(pady=15)
    
    # ========== ALERTS ==========
    
    def check_alerts(self):
        """Check for alerts on startup (low stock, expiring, overdue invoices)"""
        alerts = []
        
        try:
            conn = sqlite3.connect(self.get_database_path())
            cursor = conn.cursor()
            
            # Low stock
            cursor.execute('SELECT COUNT(*) FROM medicines WHERE quantity <= min_stock')
            low = cursor.fetchone()[0]
            if low > 0:
                alerts.append(f"⚠️ {low} medicines low in stock")
            
            # Expiring
            today = datetime.now().date()
            cursor.execute('SELECT expiry_date FROM medicines')
            expiring = 0
            for exp in cursor.fetchall():
                if exp[0] and len(exp[0]) == 10:
                    try:
                        expiry = datetime.strptime(exp[0], '%Y-%m-%d').date()
                        if 0 <= (expiry - today).days <= 30:
                            expiring += 1
                    except:
                        pass
            
            if expiring > 0:
                alerts.append(f"⏰ {expiring} medicines expiring soon")
            
            # Overdue invoices (add this block)
            cursor.execute('''
                SELECT COUNT(*) FROM invoices 
                WHERE status IN ('Pending', 'Partial') AND due_date < date('now')
            ''')
            overdue = cursor.fetchone()[0]
            if overdue > 0:
                alerts.append(f"📄 {overdue} overdue invoice(s)")
            
            conn.close()
            
        except Exception as e:
            # If invoices table doesn't exist yet, ignore
            pass
        
        if alerts:
            messagebox.showwarning("🔔 Alerts", "\n".join(alerts))
            
    def apply_settings(self):
        """Reload config, refresh sales tab, and reset dashboard authorization."""
        config = self.load_config()
        self.government_mode = config.get('government_mode', False)
        self.business_tin = config.get('tin', '')
        self._dashboard_authorized = False   # ensures dashboard asks password again
        self.refresh_sales_tab()

    def refresh_sales_tab(self):
        """Remove and recreate the sales tab, then select it."""
        for i in range(self.notebook.index('end')):
            if self.notebook.tab(i, 'text') == "💰 Sales":
                self.notebook.forget(i)
                break
        self.create_sales_tab()
        self.notebook.select(self.sales_frame)
    
    # ========== ABOUT ==========
    
    def show_about(self):
        """Professional about dialog with scrollbar"""
        about = tk.Toplevel(self.root)
        about.title("About MERawi Pharmacy Pro")
        about.geometry("550x650")
        about.configure(bg='#f5f5f5')
        about.transient(self.root)
        about.grab_set()
        
        # Center window
        about.update_idletasks()
        x = (about.winfo_screenwidth() // 2) - (550 // 2)
        y = (about.winfo_screenheight() // 2) - (650 // 2)
        about.geometry(f'+{x}+{y}')
        
        # Create canvas and scrollbar
        canvas = tk.Canvas(about, bg='#f5f5f5', highlightthickness=0)
        scrollbar = tk.Scrollbar(about, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f5f5f5')
        
        scrollable_frame.bind(
            '<Configure>',
            lambda e: canvas.configure(scrollregion=canvas.bbox('all'))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Mouse wheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), 'units')
        
        scrollable_frame.bind('<Enter>', lambda e: canvas.bind_all('<MouseWheel>', _on_mousewheel))
        scrollable_frame.bind('<Leave>', lambda e: canvas.unbind_all('<MouseWheel>'))
        
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Main content
        main = tk.Frame(scrollable_frame, bg='white', padx=30, pady=30)
        main.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Logo/Icon
        tk.Label(main, text="🏥", bg='white', font=('Segoe UI', 48)).pack(pady=(0, 10))
        
        # App name
        tk.Label(main, text="MERawi PHARMACY PRO", bg='white', fg='#2c3e50',
                font=('Segoe UI', 20, 'bold')).pack()
        
        tk.Label(main, text="Version 1.0", bg='white', fg='#7f8c8d',
                font=('Segoe UI', 11)).pack(pady=(5, 20))
        
        # Separator
        tk.Frame(main, bg='#ecf0f1', height=1).pack(fill='x', pady=10)
        
        # Description
        tk.Label(main, text="COMPLETE PHARMACY MANAGEMENT SYSTEM", bg='white', fg='#2c3e50',
                font=('Segoe UI', 12, 'bold')).pack(pady=(0, 15))
        
        tk.Label(main, text="Designed for Ethiopian pharmacies to manage inventory, sales, and business intelligence with ease.",
                bg='white', fg='#34495e', font=('Segoe UI', 10), wraplength=450, justify='center').pack(pady=(0, 20))
        
        # Features Section
        features_frame = tk.Frame(main, bg='#f8f9fa', padx=20, pady=20)
        features_frame.pack(fill='x', pady=10)
        
        tk.Label(features_frame, text="📊 KEY FEATURES", bg='#f8f9fa', fg='#2c3e50',
                font=('Segoe UI', 12, 'bold')).pack(anchor='w', pady=(0, 10))
        
        features = [
            "✓ Complete inventory management with batch tracking",
            "✓ Point of sale with tax calculation (VAT 15%)",
            "✓ Business intelligence dashboard with charts",
            "✓ Expiry date monitoring and alerts",
            "✓ Low stock notifications",
            "✓ Professional receipt printing (PDF/Text)",
            "✓ Excel import/export for data backup",
            "✓ Sales reports and analytics",
            "✓ Category-wise inventory analysis",
            "✓ Offline operation - no internet needed",
            "✓ Secure license protection",
            "✓ Automatic data backup in AppData folder"
        ]
        
        for feature in features:
            tk.Label(features_frame, text=feature, bg='#f8f9fa', fg='#34495e',
                    font=('Segoe UI', 9), anchor='w', justify='left').pack(anchor='w', pady=2)
        
        # Developer Section
        dev_frame = tk.Frame(main, bg='#f8f9fa', padx=20, pady=20)
        dev_frame.pack(fill='x', pady=10)
        
        tk.Label(dev_frame, text="👨‍💻 DEVELOPER", bg='#f8f9fa', fg='#2c3e50',
                font=('Segoe UI', 12, 'bold')).pack(anchor='w', pady=(0, 10))
        
        tk.Label(dev_frame, text="Merawi Yohannes", bg='#f8f9fa', fg='#2c3e50',
                font=('Segoe UI', 14, 'bold')).pack(anchor='w')
        
        tk.Label(dev_frame, text="Software Engineer", bg='#f8f9fa', fg='#7f8c8d',
                font=('Segoe UI', 10, 'italic')).pack(anchor='w', pady=(0, 10))
        
        # Contact Information
        contact_frame = tk.Frame(dev_frame, bg='#f8f9fa')
        contact_frame.pack(fill='x', pady=5)
        
        contacts = [
            ("📱 Phone 1:", "0921-540-245"),
            ("📱 Phone 2:", "0960-633-549"),
            ("📧 Email:", "merawiyohannes@gmail.com"),
            ("📍 Address:", "Addis Ababa, Ethiopia")
        ]
        
        for icon, value in contacts:
            row = tk.Frame(contact_frame, bg='#f8f9fa')
            row.pack(fill='x', pady=2)
            tk.Label(row, text=icon, bg='#f8f9fa', fg='#27ae60',
                    font=('Segoe UI', 9, 'bold'), width=8, anchor='w').pack(side='left')
            tk.Label(row, text=value, bg='#f8f9fa', fg='#34495e',
                    font=('Segoe UI', 9)).pack(side='left')
        
        # Support
        support_frame = tk.Frame(main, bg='#2c3e50', padx=20, pady=15)
        support_frame.pack(fill='x', pady=20)
        
        tk.Label(support_frame, text="⚕️ SUPPORT 24/7", bg='#2c3e50', fg='#f1c40f',
                font=('Segoe UI', 11, 'bold')).pack()
        tk.Label(support_frame, text="Response within 1 hour", bg='#2c3e50', fg='#ecf0f1',
                font=('Segoe UI', 9)).pack()
        
        # Copyright
        tk.Label(main, text="© 2024 MERawi Pharmacy Pro. All rights reserved.",
                bg='white', fg='#95a5a6', font=('Segoe UI', 8)).pack(pady=(10, 5))
        tk.Label(main, text="Licensed per computer - One-time payment",
                bg='white', fg='#bdc3c7', font=('Segoe UI', 7)).pack()
        
        # Close button
        close_btn = tk.Button(main, text="Close", bg='#ecf0f1', fg='#2c3e50',
                            font=('Segoe UI', 10), padx=30, pady=5,
                            relief='flat', command=about.destroy)
        close_btn.pack(pady=(20, 0))
        
        # Hover effect
        def on_enter(e): close_btn['background'] = '#d0d3d4'
        def on_leave(e): close_btn['background'] = '#ecf0f1'
        close_btn.bind('<Enter>', on_enter)
        close_btn.bind('<Leave>', on_leave)
# ========== MAIN ==========

def main():
    """Main entry point with error handling"""
    try:
        root = tk.Tk()
        app = PharmacyApp(root)
        
        if not app.check_license():
            root.quit()
            return
        
        root.mainloop()
        
    except Exception as e:
        # Show detailed error message with contact info
        error_msg = f"""❌ APPLICATION ERROR

An unexpected error occurred:

{str(e)}

Please contact the developer immediately:

👨‍💻 Merawi Yohannes
📱 0921-540-245  |  0960-633-549
📧 merawiyohannes@gmail.com

Please send a screenshot of this error message."""
        
        messagebox.showerror("🚨 Critical Error", error_msg)
        
        # Also write to log file
        try:
            log_dir = os.path.join(os.environ['LOCALAPPDATA'], 'MERawiPharmacy', 'logs')
            os.makedirs(log_dir, exist_ok=True)
            log_file = os.path.join(log_dir, f"error_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
            
            with open(log_file, 'w') as f:
                f.write(f"Error Time: {datetime.now()}\n")
                f.write(f"Error: {str(e)}\n")
                f.write(f"Type: {type(e).__name__}\n")
                import traceback
                f.write(f"Traceback:\n{traceback.format_exc()}")
            
            messagebox.showinfo("📝 Log Saved", f"Error log saved to:\n{log_file}")
            
        except:
            pass
        
        # Don't quit immediately - let user see the error
        root.mainloop()

if __name__ == "__main__":
    main()
             